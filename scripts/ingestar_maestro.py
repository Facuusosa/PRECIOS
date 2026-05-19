"""
Convierte MAESTRO.xlsx en data/raw/mapeo_brujula.json.

Estructura de salida:
  por_ean:           ean -> {sector, subcategoria, abc, nombre_verificacion, sku_yaguar?, sku_maxiconsumo?}
  por_sku_yaguar:    sku -> ean
  por_sku_maxiconsumo: sku -> ean

Uso: python scripts/ingestar_maestro.py
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no instalado. Correr: pip install openpyxl")
    sys.exit(1)

MAESTRO_PATH = Path(".claude/docs/MAESTRO.xlsx")
OUTPUT_PATH = Path("data/raw/mapeo_brujula.json")

# Correcciones de acentos para departamentos y palabras comunes en categorias
_CORRECCIONES = {
    "Almacen": "Almacén",
    "Perfumeria": "Perfumería",
    "Libreria": "Librería",
    "Ferreteria": "Ferretería",
    "Jugueteria": "Juguetería",
    "Iluminacion": "Iluminación",
    "Electricidad": "Electricidad",
    "Bebes": "Bebés",
    "Congelados": "Congelados",
    "Cosmeticos": "Cosméticos",
    "Jabones": "Jabones",
    "Panales": "Pañales",
    "Coloracion": "Coloración",
    "Depilacion": "Depilación",
    "Proteccion": "Protección",
    "Acondicionadores": "Acondicionadores",
    "Shampoo": "Shampoo",
    "Lacteos": "Lácteos",
    "Leches": "Leches",
    "Cereales": "Cereales",
    "Galletitas": "Galletitas",
    "Mermeladas": "Mermeladas",
    "Conservas": "Conservas",
    "Legumbres": "Legumbres",
    "Pastas": "Pastas",
    "Fideos": "Fideos",
    "Aceites": "Aceites",
    "Vinagres": "Vinagres",
    "Azucar": "Azúcar",
    "Azucares": "Azúcares",
    "Yerbas": "Yerbas",
    "Infusiones": "Infusiones",
    "Cafe": "Café",
    "Cafes": "Cafés",
    "Jugos": "Jugos",
    "Gaseosas": "Gaseosas",
    "Cervezas": "Cervezas",
    "Vinos": "Vinos",
    "Aperitivos": "Aperitivos",
    "Sodas": "Sodas",
    "Aguas": "Aguas",
    "Cigarrillos": "Cigarrillos",
    "Golosinas": "Golosinas",
    "Caramelos": "Caramelos",
    "Gomitas": "Gomitas",
    "Chocolates": "Chocolates",
    "Alfajores": "Alfajores",
    "Ketchup": "Ketchup",
    "Mayonesa": "Mayonesa",
    "Mostaza": "Mostaza",
    "Salsas": "Salsas",
    "Condimentos": "Condimentos",
    "Especias": "Especias",
    "Sal": "Sal",
    "Arroz": "Arroz",
    "Harinas": "Harinas",
    "Papas": "Papas",
    "Fritos": "Fritos",
    "Snacks": "Snacks",
    "Sopas": "Sopas",
    "Caldos": "Caldos",
    "Purees": "Purés",
    "Desodorantes": "Desodorantes",
    "Antitranspirantes": "Antitranspirantes",
    "Cremas": "Cremas",
    "Maquillaje": "Maquillaje",
    "Enjuagues": "Enjuagues",
    "Pastas Dentales": "Pastas Dentales",
    "Higiene": "Higiene",
    "Limpieza": "Limpieza",
    "Detergentes": "Detergentes",
    "Lavandinas": "Lavandinas",
    "Suavizantes": "Suavizantes",
    "Lustras": "Lustras",
    "Insecticidas": "Insecticidas",
    "Papeles": "Papeles",
    "Servilletas": "Servilletas",
    "Bolsas": "Bolsas",
    "Mascotas": "Mascotas",
    "Bazar": "Bazar",
    "Textil": "Textil",
    "Fiestas": "Fiestas",
    "Kiosco": "Kiosco",
    "Frescos": "Frescos",
    "Bebidas": "Bebidas",
}


def normalizar_texto(texto: str) -> str:
    """Convierte 'ALMACEN' -> 'Almacén', 'YERBAS' -> 'Yerbas'."""
    if not texto or texto == "-":
        return ""
    # Title case base
    resultado = texto.strip().title()
    # Aplicar correcciones de acentos palabra por palabra
    palabras = resultado.split()
    palabras_corregidas = [_CORRECCIONES.get(p, p) for p in palabras]
    return " ".join(palabras_corregidas)


def limpiar_nombre(nombre: str) -> str:
    """Quita '(BAJA)' y espacios extra del nombre."""
    if not nombre:
        return ""
    return re.sub(r"\s*\(BAJA\)\s*", " ", nombre).strip()


def ean_valido(valor) -> str | None:
    """Retorna el EAN como string si es valido (13 digitos numericos), sino None."""
    if not valor or valor == "-":
        return None
    ean_str = str(valor).strip().split(".")[0]  # quitar decimales si los hay
    if re.fullmatch(r"\d{13}", ean_str):
        return ean_str
    return None


def sku_valido(valor) -> str | None:
    """Retorna el SKU como string si es un numero entero positivo, sino None."""
    if not valor:
        return None
    try:
        n = int(float(str(valor)))
        return str(n) if n > 0 else None
    except (ValueError, TypeError):
        return None


def main() -> None:
    if not MAESTRO_PATH.exists():
        print(f"ERROR: No se encontro {MAESTRO_PATH}")
        sys.exit(1)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    print(f"Leyendo {MAESTRO_PATH}...")
    wb = openpyxl.load_workbook(MAESTRO_PATH, data_only=True)
    ws = wb["Sheet1"]

    # Mapear columnas por nombre
    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    COL_NOMBRE = header["Texto breve material"]
    COL_ABC    = header["Indicador ABC"]
    COL_EAN    = header["Código EAN"]
    COL_DEPTO  = header["Texto Departamento"]
    COL_CAT    = header["Texto Categoría"]
    COL_SKU_Y  = header["SKU YAGUAR"]
    COL_SKU_M  = header["SKU MAXICONSUMO"]

    mapeo = {
        "por_ean": {},
        "por_sku_yaguar": {},
        "por_sku_maxiconsumo": {},
    }

    sin_ean = 0
    procesados = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre_raw  = row[COL_NOMBRE - 1]
        abc         = row[COL_ABC - 1]
        ean_raw     = row[COL_EAN - 1]
        depto_raw   = row[COL_DEPTO - 1]
        cat_raw     = row[COL_CAT - 1]
        sku_y_raw   = row[COL_SKU_Y - 1]
        sku_m_raw   = row[COL_SKU_M - 1]

        ean = ean_valido(ean_raw)
        if not ean:
            sin_ean += 1
            continue

        nombre = limpiar_nombre(str(nombre_raw) if nombre_raw else "")
        sector = normalizar_texto(str(depto_raw) if depto_raw else "")
        subcategoria = normalizar_texto(str(cat_raw) if cat_raw else "")
        sku_y = sku_valido(sku_y_raw)
        sku_m = sku_valido(sku_m_raw)

        entrada = {
            "sector": sector,
            "subcategoria": subcategoria,
            "abc": str(abc) if abc else "",
            "nombre_verificacion": nombre,
        }
        if sku_y:
            entrada["sku_yaguar"] = sku_y
        if sku_m:
            entrada["sku_maxiconsumo"] = sku_m

        mapeo["por_ean"][ean] = entrada
        if sku_y:
            mapeo["por_sku_yaguar"][sku_y] = ean
        if sku_m:
            mapeo["por_sku_maxiconsumo"][sku_m] = ean

        procesados += 1

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(mapeo, f, ensure_ascii=False, indent=2)

    print(f"Procesados:          {procesados:,} productos")
    print(f"Sin EAN valido:      {sin_ean:,} productos (no incluidos)")
    print(f"por_ean:             {len(mapeo['por_ean']):,} entradas")
    print(f"por_sku_yaguar:      {len(mapeo['por_sku_yaguar']):,} entradas")
    print(f"por_sku_maxiconsumo: {len(mapeo['por_sku_maxiconsumo']):,} entradas")
    print(f"Guardado en:         {OUTPUT_PATH}")

    # Spot-check: Amargo Obrero
    ean_amargo = "7790950001969"
    if ean_amargo in mapeo["por_ean"]:
        print(f"\nSpot-check Amargo Obrero ({ean_amargo}):")
        import pprint
        pprint.pprint(mapeo["por_ean"][ean_amargo])
    else:
        print(f"\nAVISO: EAN {ean_amargo} no encontrado en el mapeo")


if __name__ == "__main__":
    main()
