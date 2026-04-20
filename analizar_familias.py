"""
analizar_familias.py
Detecta productos con precio sospechoso entre fuentes, identifica pack mixing,
sugiere FAMILIAs y actualiza FAMILIAS_CUSTOM.xlsx automáticamente.
"""
import json, re, os, sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl no instalado: pip install openpyxl")
    sys.exit(1)

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CATALOGO    = os.path.join(BASE_DIR, "BRUJULA-DE-PRECIOS", "data", "processed", "catalogo_unificado.json")
FAMILIAS_F  = os.path.join(BASE_DIR, "data", "raw", "FAMILIAS_CUSTOM.xlsx")

# ---------------------------------------------------------------------------
# Patrones de detección de pack/presentación en nombres de producto
# ---------------------------------------------------------------------------
# Detecta: "3u x 120g", "3 unidades", "3x120", "pack de 3", "x3", "2-pack", etc.
_PACK_RE  = re.compile(
    r"(?:^|[\s\b])(\d+)\s*(?:u(?:n(?:idades?)?)?\s*[xX]\s*\d|\s*[xX]\s*\d+\s*(?:g|ml|kg|l)\b|"
    r"\s*pack|\s*unidades?\b)|"
    r"\b(\d+)\s*[xX]\s*\d+(?:\s*g|ml|kg|l|\b)|"
    r"\b(\d+)-pack\b",
    re.I
)
# Detecta el tamaño (500ml, 120gr, 1kg, etc.)
_SIZE_RE  = re.compile(r"(\d+(?:[.,]\d+)?)\s*(g|gr|grs?|ml|kg|l|lt|lts?|cc|ul)\b", re.I)
# Normalización de unidades de tamaño
_UNIT_MAP = {"g": "g", "gr": "g", "grs": "g", "ml": "ml", "kg": "kg",
             "l": "l", "lt": "l", "lts": "l", "cc": "ml", "ul": "ml"}


def extraer_pack(nombre: str) -> int:
    """Retorna el factor de pack (3 si es 3-pack, 2 si es 2-pack, 1 si es unidad)."""
    n = nombre.lower()
    # Patrones explícitos de alta confianza
    for pat in [
        r"\b(\d+)\s*u\s*[xX]\s*\d",        # "3u x 120"
        r"\b(\d+)\s*un\s*[xX]\s*\d",       # "3un x 120"
        r"\b(\d+)\s*unidades?\b",           # "3 unidades"
        r"\b(\d+)\s*pack\b",               # "3 pack"
        r"\b(\d+)-pack\b",                 # "3-pack"
        r"\b(\d+)\s*[xX]\s*\d+\s*(?:g|ml|kg|l|gr)\b",  # "3x120g"
    ]:
        m = re.search(pat, n, re.I)
        if m:
            try:
                n_pack = int(m.group(1))
                if 2 <= n_pack <= 24:
                    return n_pack
            except (ValueError, IndexError):
                pass
    return 1


def extraer_size(nombre: str) -> str:
    """Extrae el tamaño principal (ej: '120g', '500ml')."""
    m = _SIZE_RE.search(nombre)
    if not m:
        return ""
    cantidad = m.group(1).replace(",", ".")
    unidad   = _UNIT_MAP.get(m.group(2).lower(), m.group(2).lower())
    if "." in cantidad:
        try:
            cantidad = str(int(float(cantidad)))
        except ValueError:
            pass
    return f"{cantidad}{unidad}"


def base_nombre(nombre: str) -> str:
    """Nombre sin pack count ni tamaño — para agrupar variantes."""
    n = nombre.lower()
    # Quitar pack
    n = re.sub(r"\b\d+\s*u(?:n(?:idades?)?)?\s*[xX]\s*\d+", "", n, flags=re.I)
    n = re.sub(r"\b\d+\s*[xX]\s*\d+\s*(?:g|gr|ml|kg|l)\b", "", n, flags=re.I)
    n = re.sub(r"\b\d+\s*pack\b", "", n, flags=re.I)
    n = re.sub(r"\b\d+-pack\b", "", n, flags=re.I)
    # Quitar tamaño
    n = re.sub(r"\b\d+(?:[.,]\d+)?\s*(?:g|gr|ml|kg|l|lt|cc)\b", "", n, flags=re.I)
    # Limpiar
    n = re.sub(r"[^\w\s]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def jaccard(a: str, b: str) -> float:
    wa, wb = set(a.split()), set(b.split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def generar_familia(nombre: str, pack: int, size: str) -> str:
    """Genera un nombre de FAMILIA normalizado."""
    # Extraer marca y tipo de producto del nombre (primeras palabras significativas)
    n = base_nombre(nombre)
    stop = {"de", "la", "el", "los", "las", "del", "un", "una", "x", "y", "e", "i"}
    words = [w for w in n.split() if w not in stop and len(w) > 1][:5]
    base = " ".join(words)
    if pack > 1:
        return f"{base} {pack}ux{size}" if size else f"{base} {pack}u"
    else:
        return f"{base} x{size}" if size else base


# ---------------------------------------------------------------------------
# Cargar datos
# ---------------------------------------------------------------------------
def cargar_catalogo():
    if not os.path.isfile(CATALOGO):
        print(f"[ERROR] No encontrado: {CATALOGO}")
        sys.exit(1)
    return json.load(open(CATALOGO, encoding="utf-8"))


def cargar_familias_custom():
    """Retorna (eans_ya_asignados, familias_dict, workbook)."""
    eans = set()
    fams = {}
    if not os.path.isfile(FAMILIAS_F):
        print("[WARN] FAMILIAS_CUSTOM.xlsx no encontrado")
        return eans, fams, None
    wb = openpyxl.load_workbook(FAMILIAS_F)
    if "FAMILIAS" in wb.sheetnames:
        for row in wb["FAMILIAS"].iter_rows(min_row=2, values_only=True):
            if row[0]:
                try:
                    ean = str(int(float(str(row[0])))).strip()
                except (ValueError, TypeError):
                    ean = str(row[0]).strip()
                fam = str(row[1] or "").strip().lower()
                eans.add(ean)
                fams[ean] = fam
    return eans, fams, wb


# ---------------------------------------------------------------------------
# Análisis principal
# ---------------------------------------------------------------------------
def analizar(umbral_ratio: float = 2.0, umbral_confianza: float = 0.75):
    print("=" * 60)
    print("ANALIZADOR DE FAMILIAS v1.0")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    cat       = cargar_catalogo()
    eans_ok, fams_dict, wb = cargar_familias_custom()

    # Separar: con FAMILIA conocida (fuente de verdad) vs sin FAMILIA
    con_fam  = [(p, fams_dict.get(p["id_unificado"], p.get("familia", "")))
                for p in cat if p.get("familia") or fams_dict.get(p["id_unificado"])]
    sin_fam  = [p for p in cat if not p.get("familia") and p["id_unificado"] not in eans_ok]

    # Identificar sospechosos (ratio > umbral y 2+ fuentes)
    sospechosos = []
    for p in cat:
        pvals = [v for v in p["precios"].values() if v > 0]
        if len(pvals) >= 2:
            ratio = max(pvals) / min(pvals)
            if ratio >= umbral_ratio:
                sospechosos.append((ratio, p))

    sospechosos.sort(key=lambda x: -x[0])
    nuevos_sosp = [(r, p) for r, p in sospechosos if p["id_unificado"] not in eans_ok]

    print(f"\nCatálogo: {len(cat)} productos")
    print(f"Con FAMILIA asignada: {len(eans_ok)}")
    print(f"Sospechosos ratio>{umbral_ratio}x: {len(sospechosos)} total, {len(nuevos_sosp)} sin FAMILIA aún")

    # ---------------------------------------------------------------------------
    # Detectar pares pack-vs-unidad automáticamente
    # Para cada producto sin FAMILIA y con ratio sospechoso:
    #   1. Calcular su pack y size
    #   2. Buscar en el catálogo otro producto con mismo base_nombre pero pack distinto
    #   3. Si encontramos el par -> asignar FAMILIAs a ambos
    # ---------------------------------------------------------------------------
    sugerencias = []   # (ean, familia_sugerida, confianza, motivo)

    # Índice: base_nombre -> list of products
    idx_base = defaultdict(list)
    for p in cat:
        bn = base_nombre(p["nombre_display"])
        if len(bn) > 4:
            idx_base[bn].append(p)

    # También buscar por similaridad Jaccard para bases parecidas
    bases = list(idx_base.keys())

    procesados = set()
    auto_aplicados = 0

    for ratio, p in nuevos_sosp:
        ean  = p["id_unificado"]
        if ean in procesados:
            continue

        nombre = p["nombre_display"]
        pack_p = extraer_pack(nombre)
        size_p = extraer_size(nombre)
        base_p = base_nombre(nombre)

        # Buscar pares con pack distinto
        mejor_par    = None
        mejor_sim    = 0.0
        mejor_pack_q = 1

        for base_q, lista_q in idx_base.items():
            sim = jaccard(base_p, base_q)
            if sim < umbral_confianza:
                continue
            for q in lista_q:
                if q["id_unificado"] == ean:
                    continue
                pack_q = extraer_pack(q["nombre_display"])
                if pack_q != pack_p and sim > mejor_sim:
                    mejor_sim    = sim
                    mejor_par    = q
                    mejor_pack_q = pack_q

        if mejor_par and mejor_sim >= umbral_confianza:
            fam_p = generar_familia(nombre, pack_p, size_p)
            fam_q = generar_familia(mejor_par["nombre_display"],
                                    mejor_pack_q,
                                    extraer_size(mejor_par["nombre_display"]))
            confianza = "ALTA" if mejor_sim >= 0.85 else "MEDIA"
            sugerencias.append({
                "ean":      ean,
                "nombre":   nombre,
                "familia":  fam_p,
                "confianza":confianza,
                "sim":      round(mejor_sim, 2),
                "ratio":    round(ratio, 1),
                "motivo":   f"Par detectado: pack={pack_p}u vs {mejor_pack_q}u | nombre_par: {mejor_par['nombre_display'][:40]}",
            })
            # Si el par tampoco tiene FAMILIA, sugerirle la suya
            ean_q = mejor_par["id_unificado"]
            if ean_q not in eans_ok and ean_q not in procesados:
                sugerencias.append({
                    "ean":      ean_q,
                    "nombre":   mejor_par["nombre_display"],
                    "familia":  fam_q,
                    "confianza":confianza,
                    "sim":      round(mejor_sim, 2),
                    "ratio":    0.0,
                    "motivo":   f"Par de {nombre[:40]} (pack opuesto)",
                })
            procesados.add(ean)
            procesados.add(ean_q)
        else:
            # Sin par detectado — sugerir FAMILIA solo por el nombre
            fam = generar_familia(nombre, pack_p, size_p)
            sugerencias.append({
                "ean":      ean,
                "nombre":   nombre,
                "familia":  fam,
                "confianza":"BAJA",
                "sim":      0.0,
                "ratio":    round(ratio, 1),
                "motivo":   "Sin par claro — inferido por nombre",
            })

    sugerencias_alta   = [s for s in sugerencias if s["confianza"] == "ALTA"]
    sugerencias_media  = [s for s in sugerencias if s["confianza"] == "MEDIA"]
    sugerencias_baja   = [s for s in sugerencias if s["confianza"] == "BAJA"]

    print(f"\nSugerencias generadas:")
    print(f"  ALTA confianza (auto-aplicar):  {len(sugerencias_alta)}")
    print(f"  MEDIA confianza (revisar):       {len(sugerencias_media)}")
    print(f"  BAJA confianza (revisar manual): {len(sugerencias_baja)}")

    # ---------------------------------------------------------------------------
    # Escribir al Excel
    # ---------------------------------------------------------------------------
    if not wb:
        print("\n[SKIP] No hay FAMILIAS_CUSTOM.xlsx — ejecutar primero actualizar_catalogo.py")
        return sugerencias

    ws_fam = wb["FAMILIAS"]
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    alta_fill    = PatternFill("solid", fgColor="D4EDDA")   # verde
    media_fill   = PatternFill("solid", fgColor="FFF3CD")   # amarillo
    baja_fill    = PatternFill("solid", fgColor="FFE8E8")   # rojo

    last_row = ws_fam.max_row

    agregados = 0
    for s in sugerencias:
        ean = s["ean"]
        # No duplicar
        already = False
        for row in ws_fam.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip() == ean:
                already = True
                break
        if already:
            continue

        last_row += 1
        color_fill = alta_fill if s["confianza"] == "ALTA" else (media_fill if s["confianza"] == "MEDIA" else baja_fill)
        descripcion = s["nombre"][:44]
        presentacion = f"{extraer_pack(s['nombre'])}u x {extraer_size(s['nombre'])}" if extraer_pack(s['nombre']) > 1 else f"1u x {extraer_size(s['nombre'])}"
        notas = f"[AUTO {s['confianza']}] sim={s['sim']} ratio={s['ratio']}x | {s['motivo'][:60]}"
        for col, val in enumerate([ean, s["familia"], descripcion, presentacion, notas], 1):
            cell = ws_fam.cell(last_row, col, val)
            cell.fill = color_fill
            cell.alignment = Alignment(vertical="center")
        agregados += 1

    # Actualizar hoja REVISAR con estado actual
    if "REVISAR" in wb.sheetnames:
        wb.remove(wb["REVISAR"])
    ws_rev = wb.create_sheet("REVISAR")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill    = PatternFill("solid", fgColor="1E3A5F")
    alt_fill    = PatternFill("solid", fgColor="EEF3FA")
    sosp_fill   = PatternFill("solid", fgColor="FFE8E8")

    headers = ["EAN", "NOMBRE_DISPLAY", "YAGUAR", "MAXICARREFOUR", "MAXICONSUMO",
               "RATIO_MAX", "ABC", "FAMILIA_ACTUAL", "CONFIANZA", "FAMILIA_SUGERIDA"]
    widths  = [18, 46, 12, 16, 14, 12, 6, 34, 10, 34]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws_rev.cell(1, i, h)
        c.font = hdr_fill and header_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        ws_rev.column_dimensions[get_column_letter(i)].width = w

    sugs_by_ean = {s["ean"]: s for s in sugerencias}
    all_sosp    = sorted(sospechosos, key=lambda x: -x[0])
    for r, (ratio, p) in enumerate(all_sosp, 2):
        ean   = p["id_unificado"]
        sug   = sugs_by_ean.get(ean, {})
        fill  = sosp_fill if ratio > 2.5 else (alta_fill if ean in eans_ok else alt_fill)
        row_v = [
            ean,
            p["nombre_display"],
            p["precios"].get("yaguar", 0) or "",
            p["precios"].get("maxicarrefour", 0) or "",
            p["precios"].get("maxiconsumo", 0) or "",
            round(ratio, 1),
            p.get("abc", ""),
            p.get("familia", ""),
            sug.get("confianza", "YA OK" if ean in eans_ok else ""),
            sug.get("familia", ""),
        ]
        for c, val in enumerate(row_v, 1):
            cell = ws_rev.cell(r, c, val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    ws_rev.freeze_panes = "A2"
    ws_rev.auto_filter.ref = f"A1:J{len(all_sosp)+1}"

    wb.save(FAMILIAS_F)
    print(f"\n  FAMILIAS_CUSTOM.xlsx actualizado: +{agregados} FAMILIAs agregadas")
    print(f"  Hoja REVISAR regenerada con {len(all_sosp)} sospechosos + estado de confianza")

    # ---------------------------------------------------------------------------
    # Reporte final en consola
    # ---------------------------------------------------------------------------
    if sugerencias_alta:
        print(f"\n--- ALTA confianza (auto-aplicadas) ---")
        for s in sugerencias_alta[:10]:
            print(f"  {s['ean']} | {s['nombre'][:40]} -> {s['familia']}")
        if len(sugerencias_alta) > 10:
            print(f"  ... y {len(sugerencias_alta)-10} más")

    if sugerencias_media:
        print(f"\n--- MEDIA confianza (revisar en FAMILIAS_CUSTOM.xlsx) ---")
        for s in sugerencias_media[:8]:
            print(f"  {s['ean']} | {s['nombre'][:40]} -> {s['familia']}")

    return sugerencias


if __name__ == "__main__":
    sugerencias = analizar()
    alta = sum(1 for s in sugerencias if s["confianza"] == "ALTA")
    media = sum(1 for s in sugerencias if s["confianza"] == "MEDIA")
    print(f"\nListo. Próximo paso:")
    if alta + media > 0:
        print(f"  1. Revisá FAMILIAS_CUSTOM.xlsx -> hoja REVISAR")
        print(f"  2. Corré: python actualizar_catalogo.py")
    else:
        print(f"  No hay sugerencias nuevas — catálogo limpio.")
