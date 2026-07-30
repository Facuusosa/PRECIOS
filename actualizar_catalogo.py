#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACTUALIZAR CATALOGO - Brujula de Precios  v3.0
Estrategia de matching optimizada:
  1. MaxiCarrefour como hub (100% EAN) - se procesa primero
  2. CODIGOS.xlsx mapas bidireccionales: EAN -> SKU Yaguar/Maxiconsumo
  3. Maestro como fallback por nombre normalizado
  4. Selección del scraper con MÁS productos (no el más reciente)
"""

import os, json, glob, re, unicodedata
from datetime import datetime
from collections import defaultdict
from urllib.parse import quote_plus

try:
    import openpyxl
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    print("  [WARN] openpyxl no instalado. pip install openpyxl")

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
YAGUAR_DIR      = os.path.join(BASE_DIR, "targets", "yaguar")
MAXICARRE_DIR   = os.path.join(BASE_DIR, "targets", "maxicarrefour")
MAXICONSUMO_DIR = os.path.join(BASE_DIR, "targets", "maxiconsumo")
NINI_DIR        = os.path.join(BASE_DIR, "targets", "nini")
HISTORY_DIR     = os.path.join(BASE_DIR, "data", "history")
RAW_DIR         = os.path.join(BASE_DIR, "data", "raw")
CODIGOS_FILE         = os.path.join(RAW_DIR, "CODIGOS.xlsx")
MAESTRO_FILE         = os.path.join(RAW_DIR, "Listado Maestro 09-03.xlsx")
FAMILIAS_CUSTOM_FILE = os.path.join(RAW_DIR, "FAMILIAS_CUSTOM.xlsx")
OUTPUT_FILE     = os.path.join(BASE_DIR, "BRUJULA-DE-PRECIOS", "data", "processed", "catalogo_unificado.json")

# ---------------------------------------------------------------------------
# Sectores
# ---------------------------------------------------------------------------
SECTOR_NORMALIZE = {
    "almacen": "Almacén", "almacén": "Almacén",
    "bebidas": "Bebidas", "bodega": "Bebidas",
    "frescos": "Frescos", "lacteos": "Frescos", "lácteos": "Frescos",
    "lácteos y productos frescos": "Frescos", "lacteos y productos frescos": "Frescos",
    "fiambreria": "Frescos", "fiambrería": "Frescos",
    "carniceria": "Frescos", "carnicería": "Frescos", "quesos": "Frescos",
    "limpieza": "Limpieza", "papeles e higiene": "Limpieza", "papeles": "Limpieza",
    "perfumeria": "Cuidado Personal", "perfumería": "Cuidado Personal",
    "cuidado personal": "Cuidado Personal",
    "bazar": "Bazar", "bazar y textil": "Bazar", "hogar y bazar": "Bazar",
    "productos de fiesta": "Bazar",
    "mascotas": "Mascotas",
    "kiosco": "Kiosco",
    "mundo bebe": "Bebés", "mundo bebé": "Bebés", "bebes": "Bebés", "bebés": "Bebés",
    "congelados": "Congelados",
    "desayuno y merienda": "Desayuno y Merienda", "desayuno": "Desayuno y Merienda",
    "panaderia": "Almacén", "panadería": "Almacén",
}

YAGUAR_SECTOR_MAP = {
    "almacen": "Almacén", "almacén": "Almacén",
    "bazar": "Bazar", "bebidas": "Bebidas", "bodega": "Bebidas",
    "desayuno": "Desayuno y Merienda", "frescos": "Frescos",
    "kiosco": "Kiosco", "limpieza": "Limpieza", "mascotas": "Mascotas",
    "papeles": "Limpieza", "perfumeria": "Cuidado Personal", "perfumería": "Cuidado Personal",
}

def normalizar_sector(raw):
    key = (raw or "").lower().strip()
    return SECTOR_NORMALIZE.get(key, (raw or "Almacén").strip().title())

def mapear_sector_yaguar(raw):
    key = (raw or "").lower().strip()
    for k, v in YAGUAR_SECTOR_MAP.items():
        if k in key:
            return v
    return normalizar_sector(raw)

# ---------------------------------------------------------------------------
# Normalización de nombres
# ---------------------------------------------------------------------------
def clave_nombre(nombre):
    """Clave de matching: sin acentos, unidades canónicas, sin puntuación."""
    n = (nombre or "").lower().strip()
    n = unicodedata.normalize("NFD", n)
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    # Decimales con coma → punto  ("1,5" → "1.5")
    n = re.sub(r"(\d),(\d)", r"\1.\2", n)
    # Eliminar puntuación excepto punto decimal y dígitos
    n = re.sub(r"[^a-z0-9. ]", " ", n)
    # Eliminar 'x' multiplicador antes de número ("x354ml", "x 6", "x500g" → cantidad sin x)
    n = re.sub(r"\bx\s*(\d)", r"\1", n)
    # Canonicalizar unidades de volumen/masa/cantidad
    # CM3 y CC son equivalentes a ML
    n = re.sub(r"(\d+)\s*cm3\b",  lambda m: m.group(1)+"ml",  n)
    n = re.sub(r"(\d+)\s*ccm\b",  lambda m: m.group(1)+"ml",  n)
    n = re.sub(r"(\d+)\s*cc\b",   lambda m: m.group(1)+"ml",  n)  # CC = ML
    # Litros → ml para unificar comparaciones de cantidad
    n = re.sub(r"(\d+\.?\d*)\s*lts?\b", lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+\.?\d*)\s*lt\b",   lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    # "L" aislado después de número → ml  ("1.5 l" → "1500ml", "2 l" → "2000ml")
    n = re.sub(r"(\d+\.?\d*)\s*l\b",    lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    # GRS → GR
    n = re.sub(r"(\d+)\s*grs\b", lambda m: m.group(1)+"gr", n)
    # KG → GR
    n = re.sub(r"(\d+\.?\d*)\s*kgs?\b", lambda m: str(int(float(m.group(1))*1000))+"gr", n)
    # UNIDADES: "uni" → "un" (variante larga, ej "X 2 Uni", "12 Uni")
    n = re.sub(r"(\d+)\s*uni\b", lambda m: m.group(1)+"un", n)
    # "u" suelto después de número → "un" (pack count, ej "5 u", "5u")
    n = re.sub(r"(\d+)\s*u\b", lambda m: m.group(1)+"un", n)
    # "g" suelto después de número → "gr" (gramos, ej "108 g", "500g")
    n = re.sub(r"(\d+)\s*g\b", lambda m: m.group(1)+"gr", n)
    # Pegar dígito+unidad (sin espacio) para matching exacto
    n = re.sub(r"(\d+)\s*(cc|ml|gr|kg|un|ul)\b", r"\1\2", n)
    # Eliminar puntos residuales
    n = re.sub(r"\.", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n

def normalizar_nombre_display(nombre):
    """Nombre limpio para mostrar al usuario."""
    n = (nombre or "").strip()
    if not n or len(n) < 3:
        return n
    # Eliminar X antes de cantidad
    n = re.sub(r"\bX\s*(?=\d)", "", n)
    # Title Case
    minusc = {"de","del","la","las","los","el","y","e","o","a","con","sin","en","al","por"}
    palabras = n.split()
    tc = []
    for i, p in enumerate(palabras):
        if not p: continue
        if i > 0 and p.lower() in minusc:
            tc.append(p.lower())
        else:
            tc.append(p[0].upper() + p[1:] if len(p) > 1 else p.upper())
    n = " ".join(tc)
    # Unidades canónicas (después del Title Case)
    n = re.sub(r"(\d)\s*[Cc][Cc]\b", r"\1 ml", n)
    n = re.sub(r"(\d)\s*[Ll][Tt][Ss]?\b", r"\1 L", n)
    n = re.sub(r"(\d)\s*[Mm][Ll][Ss]?\b", r"\1 ml", n)
    n = re.sub(r"(\d)\s*[Gg][Rr][Ss]?\b", r"\1 g", n)
    n = re.sub(r"(\d)\s*[Kk][Gg][Ss]?\b", r"\1 kg", n)
    n = re.sub(r"(\d)\s*G\b", r"\1 g", n)
    n = re.sub(r"(\d+)\s*[Uu][Nn]?\s*[Xx]\s*(\d+)\s*g\b", r"\1 u x \2 g", n)
    n = re.sub(r"(\d+)\s*[Uu][Nn]?\b", lambda m: m.group(1) + " u", n)
    # Eliminar abreviaciones de envase
    n = re.sub(r"\b(Pet|Bot|Pte|Sdo|Fco|Dsc|Brik)\b", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s+", " ", n).strip().strip(".")
    return n

_PLACEHOLDERS = {"/0000-", "base.png", "noimage", "placeholder", "no-image", "sin-imagen", "default"}

def _es_placeholder(url: str) -> bool:
    if not url:
        return True
    url_lower = url.lower()
    return any(p in url_lower for p in _PLACEHOLDERS)

def mejor_imagen(imagenes):
    """Prioridad: Carrefour CDN > Maxiconsumo > Yaguar. Descarta placeholders conocidos."""
    for img in imagenes:
        if not _es_placeholder(img):
            return img
    for img in imagenes:
        if img and "/0000-" not in img:
            return img
    for img in imagenes:
        if img:
            return img
    return ""

def _carrefour_search_link(ean):
    # /search/{ean} da 1 resultado exacto CUANDO el buscador conoce ese EAN. Carrefour rota
    # EANs y su buscador solo resuelve ~48%; _carrefour_links_hibrido() verifica cada uno y
    # cae a busqueda por nombre si el EAN da 0. Este es el default antes de esa verificacion.
    if not ean:
        return ""
    return f"https://comerciante.carrefour.com.ar/search/{ean}"

def _carrefour_links_hibrido(catalogo):
    # Verifica cada EAN de Carrefour contra la API del buscador. Si lo encuentra -> link
    # directo /search/{ean} (1 resultado). Si no -> busqueda por nombre (asi nunca queda
    # pantalla vacia). Robusto: ante cualquier error de red deja el /search/{ean} ya puesto.
    from urllib.parse import quote
    from concurrent.futures import ThreadPoolExecutor
    try:
        from curl_cffi import requests as _rq
    except Exception:
        return
    mc = [p for p in catalogo
          if p.get("fuentes", {}).get("maxicarrefour") and p.get("ean")]
    if not mc:
        return

    def _ok(ean):
        url = (f"https://comerciante.carrefour.com.ar/products?currentUrl=search/{ean}"
               f"&filters=&orderBy=default&currentPage=1&itemsPerPage=12&method=productsList")
        try:
            return "item_card" in _rq.get(url, impersonate="chrome120", timeout=12).text
        except Exception:
            return True  # ante duda, conservar el /search/{ean} (no romper)

    eans = [p["ean"] for p in mc]
    try:
        with ThreadPoolExecutor(max_workers=16) as ex:
            ok = dict(zip(eans, ex.map(_ok, eans)))
    except Exception:
        return
    fb = 0
    for p in mc:
        if not ok.get(p["ean"], True):
            nombre = p["fuentes"]["maxicarrefour"].get("nombre") or p.get("nombre_display") or ""
            p["fuentes"]["maxicarrefour"]["link"] = f"https://comerciante.carrefour.com.ar/search/{quote(nombre)}"
            fb += 1
    print(f"  Links Carrefour: {len(mc)-fb} EAN directo, {fb} fallback a nombre")

def _maxiconsumo_product_link(link):
    # Link directo a la ficha del producto, pero SIN el prefijo de sucursal: la URL del
    # scraper trae /sucursal_burzaco/ hardcodeado y da Forbidden si la sesion del usuario
    # no tiene esa sucursal. Sin el segmento, Magento aplica la sucursal de la sesion.
    if not link:
        return ""
    return re.sub(r"/sucursal_[^/]+/", "/", link)

def _maxiconsumo_product_link(link):
    # Link directo a la ficha del producto, pero SIN el prefijo de sucursal: la URL del
    # scraper trae /sucursal_burzaco/ hardcodeado y da Forbidden si la sesion del usuario
    # no tiene esa sucursal. Sin el segmento, Magento aplica la sucursal de la sesion.
    if not link:
        return ""
    return re.sub(r"/sucursal_[^/]+/", "/", link)

# ---------------------------------------------------------------------------
# Carga de Excel
# ---------------------------------------------------------------------------
def cargar_excel_referencia():
    """
    Retorna:
      yag_sku_to_ean  : SKU Yaguar  -> EAN
      mco_sku_to_ean  : SKU Maxiconsumo -> EAN
      nini_sku_to_ean : SKU Nini -> EAN (solo via mapeo_brujula.json -- Nini no tiene
                        hoja propia en CODIGOS.xlsx, su API no expone EAN)
      ean_to_yag_sku  : EAN -> SKU Yaguar   (para matching desde MaxiCarrefour)
      ean_to_mco_sku  : EAN -> SKU Maxiconsumo
      ean_to_nini_sku : EAN -> SKU Nini
      ean_to_master   : EAN -> {nombre, sector, categoria, abc, familia}
      nombre_norm_to_ean : nombre_normalizado -> EAN  (fallback por nombre)
      ean_to_familia  : EAN -> familia_norm   (constraint de matching de presentaciones)
    """
    yag_sku_to_ean  = {}
    mco_sku_to_ean  = {}
    nini_sku_to_ean = {}
    ean_to_yag_sku  = {}
    ean_to_mco_sku  = {}
    ean_to_nini_sku = {}
    ean_to_master   = {}
    nombre_norm_to_ean = {}
    ean_to_familia  = {}

    if not EXCEL_DISPONIBLE:
        return (yag_sku_to_ean, mco_sku_to_ean, nini_sku_to_ean,
                ean_to_yag_sku, ean_to_mco_sku, ean_to_nini_sku,
                ean_to_master, nombre_norm_to_ean, ean_to_familia)

    # --- CODIGOS.xlsx ---
    if os.path.isfile(CODIGOS_FILE):
        wb = openpyxl.load_workbook(CODIGOS_FILE, read_only=True, data_only=True)

        # YAGUAR: col1=SKU Yaguar, col2=EAN
        if "YAGUAR" in wb.sheetnames:
            for row in wb["YAGUAR"].iter_rows(min_row=2, values_only=True):
                sku_raw, ean_raw = row[1], row[2]
                if not sku_raw or not ean_raw:
                    continue
                try:
                    sku = str(int(sku_raw))
                    ean = str(int(ean_raw))
                    if len(ean) >= 8:
                        yag_sku_to_ean[sku] = ean
                        ean_to_yag_sku[ean] = sku   # mapa inverso
                except (ValueError, TypeError):
                    pass

        # MAXICONSUMO: col1=SKU, col3=EAN (Código de barras)
        if "MAXICONSUMO" in wb.sheetnames:
            for row in wb["MAXICONSUMO"].iter_rows(min_row=2, values_only=True):
                sku_raw, ean_raw = row[1], row[3]
                if not sku_raw or not ean_raw:
                    continue
                try:
                    sku = str(int(sku_raw))
                    ean = str(int(ean_raw))
                    if len(ean) >= 8:
                        mco_sku_to_ean[sku] = ean
                        ean_to_mco_sku[ean] = sku   # mapa inverso
                except (ValueError, TypeError):
                    pass

        wb.close()
        print(f"  CODIGOS.xlsx: Yaguar={len(yag_sku_to_ean)} SKUs, Maxiconsumo={len(mco_sku_to_ean)} SKUs")
        print(f"  Mapas inversos: EAN->Yaguar={len(ean_to_yag_sku)}, EAN->Maxiconsumo={len(ean_to_mco_sku)}")
    else:
        print(f"  [WARN] No encontrado: {CODIGOS_FILE}")

    # --- Listado Maestro ---
    if os.path.isfile(MAESTRO_FILE):
        wb = openpyxl.load_workbook(MAESTRO_FILE, read_only=True, data_only=True)
        ws = wb["Sheet1"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre  = row[1]
            abc     = str(row[2] or "").strip().upper()
            sector  = row[3]
            familia = row[5]
            ean_col = row[6]
            barcode = row[8]
            marca   = row[9]
            categ   = row[11]

            ean_val = None
            for v in (ean_col, barcode):
                if v and str(v).strip() not in ("-", "", "None"):
                    try:
                        ean_val = str(int(float(str(v))))
                        break
                    except (ValueError, TypeError):
                        pass

            if not ean_val or not nombre:
                continue

            nombre_str = str(nombre).strip()
            fam_str = str(familia or "").strip()
            fam_norm = fam_str.lower() if fam_str not in ("-", "", "None") else ""
            ean_to_familia[ean_val] = fam_norm
            ean_to_master[ean_val] = {
                "nombre":    normalizar_nombre_display(nombre_str),
                "sector":    normalizar_sector(str(sector or "")),
                "categoria": str(categ or familia or "").strip().title(),
                "marca":     str(marca or "").strip().title(),
                "abc":       abc,
                "familia":   fam_norm,
            }

            # Índice por nombre normalizado -> EAN (para fallback)
            clave = clave_nombre(nombre_str)
            if clave and len(clave) > 5:
                nombre_norm_to_ean[clave] = ean_val

        wb.close()
        print(f"  Listado Maestro: {len(ean_to_master)} EANs, {len(nombre_norm_to_ean)} nombres indexados")
    else:
        print(f"  [WARN] No encontrado: {MAESTRO_FILE}")

    # --- FAMILIAS_CUSTOM.xlsx (fuente propia — override sobre el Maestro) ---
    familias_custom_loaded = 0
    if os.path.isfile(FAMILIAS_CUSTOM_FILE):
        wb = openpyxl.load_workbook(FAMILIAS_CUSTOM_FILE, read_only=True, data_only=True)
        if "FAMILIAS" in wb.sheetnames:
            for row in wb["FAMILIAS"].iter_rows(min_row=2, values_only=True):
                ean_raw   = row[0]
                fam_raw   = row[1]
                if not ean_raw or not fam_raw:
                    continue
                try:
                    ean = str(int(float(str(ean_raw)))).strip()
                except (ValueError, TypeError):
                    ean = str(ean_raw).strip()
                fam = str(fam_raw).strip().lower()
                if not ean or not fam or fam in ("-", ""):
                    continue
                ean_to_familia[ean] = fam   # override: FAMILIAS_CUSTOM tiene prioridad
                if ean in ean_to_master:
                    ean_to_master[ean]["familia"] = fam
                familias_custom_loaded += 1
        wb.close()
        print(f"  FAMILIAS_CUSTOM: {familias_custom_loaded} EANs cargados (override sobre Maestro)")
    else:
        print(f"  [INFO] FAMILIAS_CUSTOM.xlsx no encontrado — usando solo Maestro para FAMILIAs")

    # --- mapeo_brujula.json (Capa 0 — matching exacto por SKU desde Maestro de Vital) ---
    mapeo_brujula_file = os.path.join(RAW_DIR, "mapeo_brujula.json")
    if os.path.isfile(mapeo_brujula_file):
        with open(mapeo_brujula_file, encoding="utf-8") as f:
            mb = json.load(f)

        nuevos_yag = nuevos_mco = nuevos_nini = nuevos_master = 0

        # Enriquecer SKU -> EAN (solo agrega, nunca sobreescribe CODIGOS.xlsx)
        for sku, ean in mb.get("por_sku_yaguar", {}).items():
            if sku not in yag_sku_to_ean:
                yag_sku_to_ean[sku] = ean
                ean_to_yag_sku.setdefault(ean, sku)
                nuevos_yag += 1

        for sku, ean in mb.get("por_sku_maxiconsumo", {}).items():
            if sku not in mco_sku_to_ean:
                mco_sku_to_ean[sku] = ean
                ean_to_mco_sku.setdefault(ean, sku)
                nuevos_mco += 1

        for sku, ean in mb.get("por_sku_nini", {}).items():
            if sku not in nini_sku_to_ean:
                nini_sku_to_ean[sku] = ean
                ean_to_nini_sku.setdefault(ean, sku)
                nuevos_nini += 1

        # Enriquecer ean_to_master con sector y subcategoria del mapeo
        for ean, datos in mb.get("por_ean", {}).items():
            if ean not in ean_to_master:
                ean_to_master[ean] = {
                    "nombre":    datos.get("nombre_verificacion", ""),
                    "sector":    datos.get("sector", ""),
                    "categoria": datos.get("subcategoria", ""),
                    "marca":     "",
                    "abc":       datos.get("abc", ""),
                    "familia":   "",
                }
                nuevos_master += 1
            else:
                # Si ya existe pero le falta sector/categoria, completar desde mapeo
                entry = ean_to_master[ean]
                if not entry.get("sector") and datos.get("sector"):
                    entry["sector"] = datos["sector"]
                if not entry.get("categoria") and datos.get("subcategoria"):
                    entry["categoria"] = datos["subcategoria"]
                if not entry.get("abc") and datos.get("abc"):
                    entry["abc"] = datos["abc"]

        print(f"  mapeo_brujula.json: +{nuevos_yag} SKUs Yaguar, +{nuevos_mco} SKUs Maxiconsumo, "
              f"+{nuevos_nini} SKUs Nini, +{nuevos_master} EANs al Maestro")
    else:
        print(f"  [INFO] mapeo_brujula.json no encontrado en {mapeo_brujula_file}")

    # Maestro dinamico (EANs de MaxiCarrefour no presentes en el estatico)
    maestro_din_file = os.path.join(RAW_DIR, "maestro_dinamico.json")
    if os.path.isfile(maestro_din_file):
        with open(maestro_din_file, encoding="utf-8") as f:
            _dm = json.load(f)
        _dm_nuevos = 0
        for clave, ean in _dm.get("por_nombre", {}).items():
            if clave not in nombre_norm_to_ean:
                nombre_norm_to_ean[clave] = ean
                _dm_nuevos += 1
        print(f"  maestro_dinamico.json: +{_dm_nuevos} nombres adicionales al indice")

    return (yag_sku_to_ean, mco_sku_to_ean, nini_sku_to_ean,
            ean_to_yag_sku, ean_to_mco_sku, ean_to_nini_sku,
            ean_to_master, nombre_norm_to_ean, ean_to_familia)

# ---------------------------------------------------------------------------
# Cantidad canónica: volumen -> ml, peso -> g. Se usa el nombre CRUDO
# (no clave_nombre) para no perder la 'L' de litros ni los decimales.
# ---------------------------------------------------------------------------
_CANT_VOL_L   = re.compile(r"(\d+(?:[.,]\d+)?)\s*(?:l|lt|lts|litro|litros)\b", re.IGNORECASE)
_CANT_VOL_ML  = re.compile(r"(\d+(?:[.,]\d+)?)\s*(?:ml|cc)\b", re.IGNORECASE)
_CANT_PESO_KG = re.compile(r"(\d+(?:[.,]\d+)?)\s*(?:kg|kilo|kilos)\b", re.IGNORECASE)
_CANT_PESO_G  = re.compile(r"(\d+(?:[.,]\d+)?)\s*(?:g|gr|grs|gramos)\b", re.IGNORECASE)

def _cantidad_canonica(nombre):
    """Devuelve (dimension, valor) -> ('vol', ml) o ('peso', g). None si no hay
    cantidad clara. Orden L->ml->kg->g para que 'kg' no caiga en 'g' ni 'ml' en 'l'."""
    s = (nombre or "").lower().replace(",", ".")
    m = _CANT_VOL_L.search(s)
    if m:  return ("vol",  round(float(m.group(1)) * 1000))
    m = _CANT_VOL_ML.search(s)
    if m:  return ("vol",  round(float(m.group(1))))
    m = _CANT_PESO_KG.search(s)
    if m:  return ("peso", round(float(m.group(1)) * 1000))
    m = _CANT_PESO_G.search(s)
    if m:  return ("peso", round(float(m.group(1))))
    return None

# ---------------------------------------------------------------------------
# Carga de scrapers
# ---------------------------------------------------------------------------
def precio_promedio(data):
    precios = [p.get("precio", 0) for p in data if p.get("precio", 0) > 0]
    return sum(precios) / len(precios) if precios else 0

def precios_validos(data):
    """Cuenta productos con precio razonable para Argentina (> $200)."""
    return sum(1 for p in data if p.get("precio", 0) > 200)

def encontrar_mejor(directorio, patron, max_check=8):
    """
    Evalua los ultimos max_check archivos y elige el mejor por recencia + score.
    Regla: si el archivo mas reciente tiene un score dentro del 5% del maximo,
    gana el mas reciente (no el que tiene 2 productos mas de un dia viejo).
    Descarta archivos con precio promedio < $200 (bug x1000).
    """
    archivos = sorted(
        glob.glob(os.path.join(directorio, patron)),
        key=os.path.getmtime, reverse=True
    )[:max_check]

    if not archivos:
        return None, []

    scores = []
    datos  = []
    for f in archivos:
        try:
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
            prom = precio_promedio(data)
            if 0 < prom < 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 200:
                        p["precio"] = round(p["precio"] * 1000, 2)
            elif prom >= 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 100:
                        p["precio"] = round(p["precio"] * 100, 2)
            scores.append(precios_validos(data))
            datos.append((f, data))
        except Exception:
            scores.append(-1)
            datos.append((f, []))

    max_score = max(scores) if scores else 0
    # Para PRECIOS, la recencia manda: un archivo fresco con menos productos vale mas
    # que uno viejo y grande (los precios cambian). Solo se descarta el mas reciente
    # si esta claramente incompleto (<70% del mejor = scraper roto / cookies a medias).
    umbral    = max_score * 0.70

    # El primer archivo de la lista ya es el mas reciente (sorted por mtime desc)
    for i, (f, data) in enumerate(datos):
        if scores[i] >= umbral:
            return f, data

    return datos[0][0], datos[0][1]


def extraer_fecha_de_timestamp(nombre):
    """De 'output_maxiconsumo_20260528_013948.json' devuelve '2026-05-28'.
    Usa la fecha del NOMBRE del archivo (mas confiable que mtime, que cambia al
    copiar). Es el fallback de frescura cuando el producto no trae fecha propia.
    Devuelve '' si no hay patron YYYYMMDD."""
    m = re.search(r"(\d{4})(\d{2})(\d{2})", os.path.basename(nombre))
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return ""


def cargar_yaguar():
    """
    Combina los últimos max_check archivos de Yaguar por SKU único.
    Para cada SKU, usa el producto del archivo más reciente con precio válido (>$200).
    Regla de frescura: SKUs que solo aparecen en archivos >30 dias se descartan —
    si Yaguar no lo scrapeo en un mes, probablemente ya no lo tiene.
    """
    from datetime import datetime as _dt
    archivos = sorted(
        glob.glob(os.path.join(YAGUAR_DIR, "output_yaguar_*.json")) +
        glob.glob(os.path.join(HISTORY_DIR, "yaguar", "output_yaguar_*.json")),
        key=os.path.getmtime, reverse=True
    )[:12]

    if not archivos:
        print("  [SKIP] No se encontró output de Yaguar")
        return []

    hoy_ts = _dt.now().timestamp()
    LIMITE_DIAS = 30
    LIMITE_TS   = hoy_ts - LIMITE_DIAS * 86400

    sku_to_mejor  = {}
    sku_tiene_reciente = set()  # SKUs vistos en archivos de <= 30 dias
    archivos_validos = 0

    for f in archivos:
        try:
            es_reciente = os.path.getmtime(f) >= LIMITE_TS
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
            if not data:
                continue
            fecha_archivo = extraer_fecha_de_timestamp(f)
            prom = precio_promedio(data)
            if 0 < prom < 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 200:
                        p["precio"] = round(p["precio"] * 1000, 2)
            elif prom >= 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 100:
                        p["precio"] = round(p["precio"] * 100, 2)
            archivos_validos += 1
            for p in data:
                sku = str(p.get("sku", "")).strip()
                precio = p.get("precio", 0)
                if not sku or precio <= 0:
                    continue
                if not p.get("fecha"):
                    p["fecha"] = fecha_archivo
                if es_reciente:
                    sku_tiene_reciente.add(sku)
                existing = sku_to_mejor.get(sku)
                if existing is None:
                    sku_to_mejor[sku] = p
                else:
                    precio_ex = existing.get("precio", 0)
                    if precio > 200 and precio_ex < 200:
                        sku_to_mejor[sku] = p
                    elif precio > 200 and precio_ex > 200:
                        # Ambos válidos — existing es más reciente (archivos ordenados desc)
                        # Si el más antiguo (p) tiene imagen real y el más reciente no, heredar imagen
                        def _img_real(prod):
                            img = prod.get("imagen", "")
                            return bool(img and "base.png" not in img and not img.startswith("data:"))
                        if _img_real(p) and not _img_real(existing):
                            existing["imagen"] = p["imagen"]
                        # Si solo p tiene link/imagen (y existing no tiene nada útil), reemplazar
                        tiene_datos = bool(p.get("link") or p.get("imagen"))
                        existing_tiene_datos = bool(existing.get("link") or existing.get("imagen"))
                        if tiene_datos and not existing_tiene_datos:
                            sku_to_mejor[sku] = p
        except Exception:
            pass

    # Descartar SKUs que solo aparecen en archivos viejos (>30 dias)
    skus_descartados = [s for s in sku_to_mejor if s not in sku_tiene_reciente]
    for s in skus_descartados:
        del sku_to_mejor[s]

    combined = list(sku_to_mejor.values())
    con_precio = precios_validos(combined)
    if skus_descartados:
        print(f"  Yaguar: {archivos_validos} archivos combinados -> {len(combined)} prods únicos ({con_precio} válidos) | {len(skus_descartados)} SKUs viejos descartados (>30 dias sin aparecer)")
    else:
        print(f"  Yaguar: {archivos_validos} archivos combinados -> {len(combined)} prods únicos ({con_precio} válidos)")
    return combined


def _fallback_mc_desde_catalogo():
    # Cuando el scraper de MC falla en Railway (container efimero, sin outputs previos),
    # lee el catalogo_unificado.json ya commiteado y reconstruye los productos de MC.
    # CRITICO: conservar la fecha_scraping ORIGINAL. Pisar la fecha con "hoy" creaba
    # un bucle de reciclaje invisible — precios de hace 14 dias mostrados como frescos
    # (detectado 11/06: 5.128 precios MC identicos al scraping del 28/05).
    catalogo_path = os.path.join(
        BASE_DIR, "BRUJULA-DE-PRECIOS", "data", "processed", "catalogo_unificado.json"
    )
    if not os.path.exists(catalogo_path):
        print("  [MC-FALLBACK] No hay catalogo previo disponible - MC queda vacio")
        return []
    try:
        with open(catalogo_path, encoding="utf-8") as f:
            catalogo = json.load(f)
    except Exception as e:
        print(f"  [MC-FALLBACK] Error leyendo catalogo previo: {e}")
        return []

    reconstruidos = []
    fecha_original = ""
    for prod in catalogo:
        precio_mc = prod.get("precios", {}).get("maxicarrefour", 0)
        if not precio_mc or precio_mc <= 0:
            continue
        fuente = prod.get("fuentes", {}).get("maxicarrefour", {})
        fecha_fuente = fuente.get("fecha_scraping", "")
        if not fecha_original:
            fecha_original = fecha_fuente or "?"
        reconstruidos.append({
            "ean":             prod.get("ean") or prod.get("id_unificado", ""),
            "nombre":          fuente.get("nombre") or prod.get("nombre_display", ""),
            "precio":          precio_mc,
            "precio_regular":  fuente.get("precio_regular", 0),
            "oferta":          fuente.get("oferta", ""),
            "link":            fuente.get("link", ""),
            "imagen":          fuente.get("imagen", prod.get("imagen", "")),
            "sector":          prod.get("sector", ""),
            "fecha_scraping":  fecha_fuente,
            "fallback_fuente": fecha_original,
            "stock":           True,
        })

    print(f"  [MC-FALLBACK] ATENCION: scraper MC fallo - reciclando {len(reconstruidos)} "
          f"precios del catalogo anterior (fecha real: {fecha_original}). "
          f"Si esto se repite varios dias, renovar cookies MC en Railway.")
    return reconstruidos


def cargar_maxicarrefour():
    archivo, data = encontrar_mejor(MAXICARRE_DIR, "output_maxicarrefour_*.json")
    if not archivo:
        print("  [SKIP] No se encontro output de MaxiCarrefour - activando fallback")
        return _fallback_mc_desde_catalogo()
    con_precio = sum(1 for p in data if p.get("precio", 0) > 0)
    print(f"  MaxiCarrefour: {os.path.basename(archivo)} -> {len(data)} productos ({con_precio} con precio)")
    if con_precio == 0:
        print("  [WARN] Todos los precios son 0 - cookies vencidas - activando fallback")
        return _fallback_mc_desde_catalogo()
    return data


def cargar_maxiconsumo():
    """
    Carga y combina los archivos de Maxiconsumo disponibles.
    Estrategia: para cada SKU, prefiere el precio del archivo con mayor
    promedio (enriquecido) si es válido (>$200); sino usa el raw.
    Así se aprovecha la mayor cobertura del raw y la mayor calidad del enriquecido.
    """
    if not os.path.isdir(MAXICONSUMO_DIR):
        return []

    archivos = sorted(
        glob.glob(os.path.join(MAXICONSUMO_DIR, "output_maxiconsumo_*.json")) +
        glob.glob(os.path.join(HISTORY_DIR, "maxiconsumo", "output_maxiconsumo_*.json")),
        key=os.path.getmtime, reverse=True
    )[:8]
    if not archivos:
        return []

    # Cargar todos los archivos y construir un mapa SKU → mejor precio
    sku_to_mejor = {}   # sku → producto con mejor precio validado
    archivos_cargados = []

    for f in archivos:
        try:
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
            fecha_archivo = extraer_fecha_de_timestamp(f)
            for p in data:
                if not p.get("fecha_scraping"):
                    p["fecha_scraping"] = fecha_archivo
            prom = precio_promedio(data)
            # Fix precio × 1000 si el promedio es sospechosamente bajo
            if 0 < prom < 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 200:
                        p["precio"] = round(p["precio"] * 1000, 2)
            elif prom >= 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 100:
                        p["precio"] = round(p["precio"] * 100, 2)
            archivos_cargados.append((f, data, precio_promedio(data)))
        except Exception:
            pass

    if not archivos_cargados:
        return []

    # Combinar: para cada SKU usar el archivo con precio más alto (>$200 preferido)
    for _, data, prom_f in archivos_cargados:
        for p in data:
            sku = str(p.get("sku", "")).strip()
            precio = p.get("precio", 0)
            if not sku or precio <= 0:
                continue
            existing = sku_to_mejor.get(sku)
            if existing is None:
                sku_to_mejor[sku] = p
            else:
                precio_ex = existing.get("precio", 0)
                # Preferir precio válido (>$200) sobre inválido (<$200)
                if precio > 200 and precio_ex < 200:
                    sku_to_mejor[sku] = p
                elif precio > 200 and precio_ex > 200 and precio > precio_ex:
                    # Si ambos válidos, quedarse con el del archivo más reciente
                    # (el primero en la iteración ya es el más reciente, no sobreescribir)
                    pass

    combined = list(sku_to_mejor.values())
    con_precio = sum(1 for p in combined if p.get("precio", 0) > 200)
    bajos = sum(1 for p in combined if 0 < p.get("precio", 0) < 200)
    print(f"  Maxiconsumo: {len(archivos_cargados)} archivos combinados -> {len(combined)} productos")
    print(f"    {con_precio} precios válidos (>$200), {bajos} precios bajos (<$200)")
    return combined


def cargar_nini():
    """
    Combina los últimos archivos de Nini por SKU único (mismo criterio de
    frescura que cargar_yaguar(): SKUs vistos solo en archivos >30 dias se
    descartan). La API de Nini ya filtra withStock=true y entrega precio
    numerico limpio -- no hace falta el fix x1000/x100 de otros scrapers.
    """
    from datetime import datetime as _dt
    archivos = sorted(
        glob.glob(os.path.join(NINI_DIR, "output_nini_*.json")) +
        glob.glob(os.path.join(HISTORY_DIR, "nini", "output_nini_*.json")),
        key=os.path.getmtime, reverse=True
    )[:12]

    if not archivos:
        print("  [SKIP] No se encontró output de Nini")
        return []

    hoy_ts = _dt.now().timestamp()
    LIMITE_DIAS = 30
    LIMITE_TS   = hoy_ts - LIMITE_DIAS * 86400

    sku_to_mejor = {}
    sku_tiene_reciente = set()
    archivos_validos = 0

    for f in archivos:
        try:
            es_reciente = os.path.getmtime(f) >= LIMITE_TS
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
            if not data:
                continue
            fecha_archivo = extraer_fecha_de_timestamp(f)
            archivos_validos += 1
            for p in data:
                sku = str(p.get("sku", "")).strip()
                precio = p.get("precio", 0)
                if not sku or precio <= 0:
                    continue
                if not p.get("fecha"):
                    p["fecha"] = fecha_archivo
                if es_reciente:
                    sku_tiene_reciente.add(sku)
                existing = sku_to_mejor.get(sku)
                if existing is None or precio > existing.get("precio", 0):
                    sku_to_mejor[sku] = p
        except Exception:
            pass

    skus_descartados = [s for s in sku_to_mejor if s not in sku_tiene_reciente]
    for s in skus_descartados:
        del sku_to_mejor[s]

    combined = list(sku_to_mejor.values())
    con_precio = precios_validos(combined)
    if skus_descartados:
        print(f"  Nini: {archivos_validos} archivos combinados -> {len(combined)} prods únicos ({con_precio} válidos) | {len(skus_descartados)} SKUs viejos descartados (>30 dias sin aparecer)")
    else:
        print(f"  Nini: {archivos_validos} archivos combinados -> {len(combined)} prods únicos ({con_precio} válidos)")
    return combined


def cargar_cadena(clave, etiqueta):
    """
    Carga generica de fuentes tipo "cadena" (Coto, Carrefour retail):
    combina los output_{clave}_*.json recientes (targets/{clave} +
    data/history/{clave}), dedupe por EAN quedandose con el dato del archivo
    mas reciente. Las cadenas son 100% EAN — sin fuzzy.
    """
    dir_targets = os.path.join(BASE_DIR, "targets", clave)
    archivos = sorted(
        glob.glob(os.path.join(dir_targets, f"output_{clave}_*.json")) +
        glob.glob(os.path.join(HISTORY_DIR, clave, f"output_{clave}_*.json")),
        key=os.path.getmtime, reverse=True
    )[:3]
    if not archivos:
        print(f"  [SKIP] No se encontro output de {etiqueta} - catalogo sin su precio gondola")
        return []

    por_ean = {}
    for f in archivos:  # el mas reciente primero: su dato gana
        try:
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
        except Exception as e:
            print(f"  [WARN] {etiqueta}: error leyendo {os.path.basename(f)}: {e}")
            continue
        fecha_archivo = extraer_fecha_de_timestamp(f)
        for p in data:
            ean = str(p.get("ean", "")).strip()
            if not ean or p.get("precio", 0) <= 0:
                continue
            if not p.get("fecha_scraping"):
                p["fecha_scraping"] = fecha_archivo
            if ean not in por_ean:
                por_ean[ean] = p

    combined = list(por_ean.values())
    print(f"  {etiqueta}: {len(archivos)} archivo(s) -> {len(combined)} productos con EAN y precio")
    return combined


def cargar_hunterprice():
    ruta = os.path.join(BASE_DIR, "archive", "data_hunterprice.json")
    if not os.path.isfile(ruta):
        print("  [SKIP] No encontrado: archive/data_hunterprice.json")
        return []
    with open(ruta, encoding="utf-8") as f:
        data = json.load(f)
    # Solo los que tienen MaxiCarrefour (nuestro hub EAN) como referencia
    filtrado = [p for p in data if p.get("MAXI CARREFOUR") and p["MAXI CARREFOUR"] > 0]
    print(f"  Hunterprice: {len(data)} productos total, {len(filtrado)} con precio MaxiCarrefour")
    return filtrado

# ---------------------------------------------------------------------------
# Constructor del catálogo unificado
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Expansión autónoma del mapeo SKU->EAN usando fuentes con EAN nativo
# (cadenas + MaxiCarrefour) como diccionario de nombres.
#   - sim >= 0.85 -> se aprende solo y se persiste en mapeo_brujula.json
#   - 0.75 <= sim < 0.85 -> cola de revisión en data/quality/matches_pendientes.json
# Guards anti falso positivo: marca (primer token significativo) igual +
# cantidad canónica en la misma dimensión con tolerancia 10% (misma idea que 6d).
# Diseñado 14/07/2026: simulación midió 133 matches Yaguar + 34 MCO en banda
# alta con 12/12 correctos en muestra; el índice se realimenta con cada scrape.
# ---------------------------------------------------------------------------
MAPEO_BRUJULA_FILE      = os.path.join(RAW_DIR, "mapeo_brujula.json")
MATCHES_PENDIENTES_FILE = os.path.join(BASE_DIR, "data", "quality", "matches_pendientes.json")
MAPEOS_SOSPECHOSOS_FILE = os.path.join(BASE_DIR, "data", "quality", "mapeos_sospechosos.json")

# ---------------------------------------------------------------------------
# Alias de EAN curados a mano (data/quality/alias_ean.json): mismo producto
# fisico con 2 codigos de barra (caso Fernet Branca 750, 16/07/2026 — el precio
# MCO quedaba en una entrada duplicada y la comparativa se partia en dos).
# El reemplazo corre en cada punto donde una fuente resuelve su EAN, ANTES de
# agrupar — asi la maquinaria existente unifica sin fuzzy. SOLO entran pares
# aprobados por Facu; nada automatico escribe ese archivo (regla 09).
# ---------------------------------------------------------------------------
ALIAS_EAN_FILE = os.path.join(BASE_DIR, "data", "quality", "alias_ean.json")

def _cargar_alias_ean():
    try:
        with open(ALIAS_EAN_FILE, encoding="utf-8") as f:
            alias = json.load(f).get("alias", {})
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
    # Resolver cadenas de 1 nivel (a->b, b->c => a->c) y evitar ciclos
    return {a: alias.get(c, c) for a, c in alias.items() if a != alias.get(c, c)}

_ALIAS_EAN = _cargar_alias_ean()

def _ean_canonico(ean):
    return _ALIAS_EAN.get(ean, ean)

_EXP_TH_AUTO     = 0.85
_EXP_TH_REVISION = 0.75
_EXP_STOP = {"de", "la", "el", "en", "y", "x", "con", "por", "para", "un", "una",
             "del", "los", "las", "al", "ml", "gr", "cc", "kg", "sin", "bot",
             "pet", "brik", "fco", "sdo", "pvc", "p", "s", "c"}
# Tokens de categoría que no identifican marca (para elegir el token de marca)
_EXP_CATEGORIA = {"gaseosa", "agua", "soda", "jugo", "cerveza", "vino", "aceite",
                  "arroz", "azucar", "harina", "leche", "yogur", "manteca", "queso",
                  "yerba", "cafe", "te", "sal", "fideos", "galletita", "galleta",
                  "chocolate", "caramelo", "caramelos", "golosina", "snack", "papel",
                  "detergente", "jabon", "shampoo", "desodorante", "lavandina",
                  "suavizante", "servilleta", "mayonesa", "ketchup", "mostaza",
                  "salsa", "vinagre", "gelatina", "polvo", "mate", "cocido",
                  "licor", "amargo", "atun", "pure", "edulcorante", "espuma",
                  "polenta", "desinfectante", "higienico", "hig",
                  "parcialmente", "descremada", "entera", "liviana"}

# Abreviaturas que las cadenas usan y los mayoristas escriben completas (o al
# reves). Caso leche 14/07: Carrefour "ultra parc descremada multivit" vs Yaguar
# "PARCIALMENTE DESCREMADA" daba sim 0.67 (<0.75) y el match no se proponia.
# Solo expansiones univocas — nada ambiguo tipo "desc" (descremada/descartable).
_EXP_ABREV = {
    "parc": "parcialmente", "parcial": "parcialmente",
    "descr": "descremada", "descrem": "descremada",
    "past": "pasteurizada",
    "multivit": "multivitaminas", "multivitaminica": "multivitaminas",
    "ls": "serenisima",   # el Maestro abrevia "Leche LS Liviana..."
}

def _exp_tokens(nombre):
    # Los digitos sueltos SE CONSERVAN: "1"/"2"/"3" suelen ser el % de grasa u
    # otra variante (caso 14/07: leche 2% se llevo el EAN de la 1% porque ambos
    # tokenizaban identico). Un digito decorativo baja un poco el sim de pares
    # buenos (falso negativo leve) — preferible a fusionar variantes (grave).
    cl = clave_nombre(nombre)
    return {_EXP_ABREV.get(w, w) for w in cl.split()
            if (len(w) > 1 or w.isdigit()) and w not in _EXP_STOP}

def _exp_marca(nombre):
    cl = clave_nombre(nombre)
    for w in cl.split():
        if len(w) > 2 and w not in _EXP_STOP and w not in _EXP_CATEGORIA and not w[0].isdigit():
            return w
    return ""

def _tfidf_trigramas(textos):
    """Matriz TF-IDF sparse de trigramas de caracteres, normalizada L2.
    Hecha a mano con scipy para no sumar scikit-learn (~80MB) al pipeline.
    Devuelve (matriz_csr, vocabulario) — vocabulario para vectorizar queries."""
    from scipy.sparse import csr_matrix
    import math
    vocab = {}
    filas, cols, vals = [], [], []
    df = defaultdict(int)
    docs_tri = []
    for texto in textos:
        s = f" {texto} "
        tri = defaultdict(int)
        for i in range(len(s) - 2):
            tri[s[i:i+3]] += 1
        docs_tri.append(tri)
        for g in tri:
            df[g] += 1
    n = len(textos)
    idf = {g: math.log(n / (1 + d)) + 1 for g, d in df.items()}
    for fi, tri in enumerate(docs_tri):
        norm = 0.0
        pares = []
        for g, tf in tri.items():
            if g not in vocab:
                vocab[g] = len(vocab)
            w = tf * idf[g]
            pares.append((vocab[g], w))
            norm += w * w
        norm = math.sqrt(norm) or 1.0
        for ci, w in pares:
            filas.append(fi); cols.append(ci); vals.append(w / norm)
    m = csr_matrix((vals, (filas, cols)), shape=(n, len(vocab)))
    return m, vocab, idf


def expandir_mapeo_con_cadenas(yaguar_data, maxiconsumo_data, fuentes_con_ean,
                               yag_sku_to_ean, mco_sku_to_ean,
                               ean_to_yag_sku, ean_to_mco_sku,
                               nombre_norm_to_ean, ean_to_master,
                               nini_data=None, nini_sku_to_ean=None, ean_to_nini_sku=None):
    """Resuelve EANs de productos Yaguar/Maxiconsumo que CODIGOS y el Maestro no
    cubren, matcheando sus nombres contra los pares EAN+nombre frescos de las
    fuentes con EAN nativo. MUTA los dicts de mapeo en memoria (la corrida actual
    ya aprovecha lo aprendido) y persiste el aprendizaje en mapeo_brujula.json.
    fuentes_con_ean: lista de (nombre_fuente, data) con campos ean/nombre/precio.
    """
    nini_data = nini_data or []
    nini_sku_to_ean = nini_sku_to_ean if nini_sku_to_ean is not None else {}
    ean_to_nini_sku = ean_to_nini_sku if ean_to_nini_sku is not None else {}

    # --- 1. Índice EAN -> nombres de las fuentes confiables ---
    entries  = []                  # (ean, fuente, token_set, marca, qty, nombre)
    word_idx = defaultdict(list)
    ean_nombres = defaultdict(dict)   # ean -> {fuente: nombre} (para auto-correccion)

    def _indexar(ean, fuente, nombre):
        ean_nombres[ean][fuente] = nombre
        ws = _exp_tokens(nombre)
        if not ws:
            return
        i = len(entries)
        entries.append((ean, fuente, ws, _exp_marca(nombre), _cantidad_canonica(nombre), nombre))
        for w in ws:
            word_idx[w].append(i)

    for fuente, data in fuentes_con_ean:
        for p in data:
            ean = str(p.get("ean", "")).strip()
            nombre = p.get("nombre", "")
            if not ean or len(ean) < 8 or not nombre or p.get("precio", 0) <= 0:
                continue
            _indexar(ean, fuente, nombre)

    # El Listado Maestro tambien entra al indice (pedido Facu 15/07/2026): sus
    # 25k nombres ya alimentaban el lookup exacto y el fuzzy 1b, pero no esta
    # via con guards nuevos (marca en cualquier posicion, digitos, bandas).
    for ean, m in ean_to_master.items():
        nombre = (m or {}).get("nombre", "")
        if nombre and len(nombre) >= 4 and len(ean) >= 8 and ean not in ean_nombres:
            _indexar(ean, "maestro", nombre)

    def _buscar(nombre):
        ws_p = _exp_tokens(nombre)
        if not ws_p:
            return None
        m_p = _exp_marca(nombre)
        q_p = _cantidad_canonica(nombre)
        cands = set()
        for w in ws_p:
            cands.update(word_idx.get(w, []))
        best_sim, best_i = 0.0, None
        for i in cands:
            ean, fte, ws_c, m_c, q_c, nom_c = entries[i]
            sim = len(ws_p & ws_c) / len(ws_p | ws_c)
            if sim <= best_sim:
                continue
            if m_p and m_c and m_p != m_c:
                # El "primer token significativo" no siempre es la marca (Yaguar:
                # "PARCIALMENTE DESCREMADA LA SERENISIMA..." da "parcialmente").
                # Aceptar si la marca detectada de un lado esta como token en el
                # otro; rechazar solo si ninguna aparece enfrente (MANAOS/IVESS).
                if m_c not in ws_p and m_p not in ws_c:
                    continue
            if q_p and q_c:
                if q_p[0] != q_c[0]:
                    continue
                if abs(q_p[1] - q_c[1]) > 0.10 * max(q_p[1], q_c[1]):
                    continue
            best_sim, best_i = sim, i
        return (best_sim, best_i) if best_i is not None else None

    # --- 2. Cargar estado persistente ---
    mb = {}
    if os.path.isfile(MAPEO_BRUJULA_FILE):
        with open(MAPEO_BRUJULA_FILE, encoding="utf-8") as f:
            mb = json.load(f)
    mb.setdefault("por_sku_yaguar", {})
    mb.setdefault("por_sku_maxiconsumo", {})
    mb.setdefault("por_sku_nini", {})
    mb.setdefault("por_ean", {})

    rechazados = set()
    if os.path.isfile(MATCHES_PENDIENTES_FILE):
        try:
            with open(MATCHES_PENDIENTES_FILE, encoding="utf-8") as f:
                rechazados = set(json.load(f).get("rechazados", []))
        except (json.JSONDecodeError, OSError):
            pass

    # --- 3. Buscar matches para los productos sin EAN ---
    pendientes = []
    sin_match = []   # (etiqueta, sku, nombre, tokens) -> pase TF-IDF 3b
    n_auto = {"yaguar": 0, "maxiconsumo": 0, "nini": 0}
    lotes = [
        ("yaguar",      yaguar_data,      yag_sku_to_ean,  ean_to_yag_sku,  "por_sku_yaguar"),
        ("maxiconsumo", maxiconsumo_data, mco_sku_to_ean,  ean_to_mco_sku,  "por_sku_maxiconsumo"),
        ("nini",        nini_data,        nini_sku_to_ean, ean_to_nini_sku, "por_sku_nini"),
    ]
    for etiqueta, data, sku_to_ean, ean_to_sku, clave_mb in lotes:
        for p in data:
            sku = str(p.get("sku", "")).strip()
            nombre = p.get("nombre", "")
            if not sku or not nombre or p.get("precio", 0) <= 0:
                continue
            if sku_to_ean.get(sku) or nombre_norm_to_ean.get(clave_nombre(nombre)):
                continue  # ya resuelve por CODIGOS/mapeo/Maestro
            r = _buscar(nombre)
            if not r:
                sin_match.append((etiqueta, sku, nombre, _exp_tokens(nombre)))
                continue
            sim, i = r
            if sim < _EXP_TH_REVISION:
                sin_match.append((etiqueta, sku, nombre, _exp_tokens(nombre)))
            ean, fte, _, _, _, nom_cadena = entries[i]
            if f"{etiqueta}:{sku}:{ean}" in rechazados:
                continue  # rechazado a mano por Facu — no re-aprender nunca
            if sim >= _EXP_TH_AUTO:
                if ean in ean_to_sku:
                    continue  # ese EAN ya pertenece a otro SKU de esta fuente
                sku_to_ean[sku] = ean
                ean_to_sku[ean] = sku
                mb[clave_mb][sku] = ean
                # Display limpio: si el Maestro no conoce el EAN, usar el nombre
                # de la cadena como nombre de verificación
                if ean not in ean_to_master:
                    ean_to_master[ean] = {"nombre": nom_cadena, "sector": "",
                                          "categoria": "", "marca": "", "abc": "",
                                          "familia": ""}
                    mb["por_ean"].setdefault(ean, {"nombre_verificacion": nom_cadena})
                n_auto[etiqueta] += 1
            elif sim >= _EXP_TH_REVISION:
                clave_rechazo = f"{etiqueta}:{sku}:{ean}"
                if clave_rechazo not in rechazados:
                    pendientes.append({
                        "fuente": etiqueta, "sku": sku, "nombre": nombre,
                        "ean_candidato": ean, "nombre_cadena": nom_cadena,
                        "fuente_cadena": fte, "similitud": round(sim, 3),
                    })

    # --- 3b. Pase TF-IDF de trigramas — MODO CALIBRACION: solo propone ---
    # Encuentra candidatos que el Jaccard de palabras enteras no ve (abreviaturas
    # sin mapear, typos, reordenamientos). NO auto-aprende: todo va a pendientes
    # hasta calibrar el umbral con el feedback de Facu sobre la primera tanda.
    _TFIDF_TH = 0.80
    if sin_match and entries:
        try:
            from scipy.sparse import csr_matrix
            import math as _math
            corpus = [" ".join(sorted(ws)) for _, _, ws, _, _, _ in entries]
            m_idx, vocab, idf = _tfidf_trigramas(corpus)
            m_idx_t = m_idx.T.tocsr()
            tf_prop = 0
            B = 200
            for b0 in range(0, len(sin_match), B):
                bloque = sin_match[b0:b0 + B]
                filas, cols, vals = [], [], []
                for qi, (_, _, _, ws_q) in enumerate(bloque):
                    s = f" {' '.join(sorted(ws_q))} "
                    tri = defaultdict(int)
                    for i in range(len(s) - 2):
                        tri[s[i:i+3]] += 1
                    norm, pares = 0.0, []
                    for g, tf in tri.items():
                        if g in vocab:
                            w = tf * idf[g]
                            pares.append((vocab[g], w)); norm += w * w
                    norm = _math.sqrt(norm) or 1.0
                    for ci, w in pares:
                        filas.append(qi); cols.append(ci); vals.append(w / norm)
                q = csr_matrix((vals, (filas, cols)), shape=(len(bloque), len(vocab)))
                sims = q.dot(m_idx_t).tocsr()
                for qi in range(len(bloque)):
                    fila = sims.getrow(qi)
                    if fila.nnz == 0:
                        continue
                    j = int(fila.indices[fila.data.argmax()])
                    cos = float(fila.data.max())
                    if cos < _TFIDF_TH:
                        continue
                    etiqueta, sku, nombre_q, ws_q = bloque[qi]
                    ean, fte, ws_c, m_c, q_c, nom_c = entries[j]
                    # mismos guards que el pase principal
                    if f"{etiqueta}:{sku}:{ean}" in rechazados:
                        continue
                    m_q = _exp_marca(nombre_q)
                    if m_q and m_c and m_q != m_c and m_c not in ws_q and m_q not in ws_c:
                        continue
                    q_q = _cantidad_canonica(nombre_q)
                    if q_q and q_c:
                        if q_q[0] != q_c[0]:
                            continue
                        if abs(q_q[1] - q_c[1]) > 0.10 * max(q_q[1], q_c[1]):
                            continue
                    d_q = {t for t in ws_q if t.isdigit()}
                    d_c = {t for t in ws_c if t.isdigit()}
                    if d_q and d_c and not (d_q & d_c):
                        continue
                    pendientes.append({"fuente": etiqueta, "sku": sku, "nombre": nombre_q,
                                       "ean_candidato": ean, "nombre_cadena": nom_c,
                                       "fuente_cadena": fte,
                                       "similitud": round(cos, 3), "via": "tfidf"})
                    tf_prop += 1
            print(f"  TF-IDF trigramas: {tf_prop} candidatos nuevos (solo revision — modo calibracion)")
        except Exception as e:
            print(f"  [WARN] Pase TF-IDF fallo (se continua sin el): {e}")

    # --- 4. Auto-corrección: auditar lo aprendido contra los nombres frescos ---
    # Solo los mapeos APRENDIDOS (mapeo_brujula) — CODIGOS.xlsx es curado a mano.
    # Los packs NxM se saltean: "4UNX276GR" (total) vs "4 Uni X 69 Gr" (unitario)
    # es el mismo producto expresado distinto y genera falsa alarma.
    # Los que Facu ya reviso y mantuvo (lista 'aceptados') no se re-proponen.
    _pack_re = re.compile(r"\d+\s*(?:un|uni|u)?\s*x\s*\d+", re.IGNORECASE)
    aceptados = set()
    if os.path.isfile(MAPEOS_SOSPECHOSOS_FILE):
        try:
            with open(MAPEOS_SOSPECHOSOS_FILE, encoding="utf-8") as f:
                aceptados = set(json.load(f).get("aceptados", []))
        except (json.JSONDecodeError, OSError):
            pass
    sospechosos = []
    por_sku_data = {"por_sku_yaguar": {str(p.get("sku", "")).strip(): p.get("nombre", "")
                                       for p in yaguar_data},
                    "por_sku_maxiconsumo": {str(p.get("sku", "")).strip(): p.get("nombre", "")
                                            for p in maxiconsumo_data},
                    "por_sku_nini": {str(p.get("sku", "")).strip(): p.get("nombre", "")
                                     for p in nini_data}}
    for clave_mb, sku_nombres in por_sku_data.items():
        for sku, ean in mb.get(clave_mb, {}).items():
            nombre_may = sku_nombres.get(sku, "")
            if not nombre_may or ean not in ean_nombres:
                continue
            if f"{clave_mb}:{sku}:{ean}" in aceptados:
                continue
            if _pack_re.search(nombre_may):
                continue
            q_may = _cantidad_canonica(nombre_may)
            if not q_may:
                continue
            for fte, nom_cad in ean_nombres[ean].items():
                if _pack_re.search(nom_cad):
                    continue
                q_cad = _cantidad_canonica(nom_cad)
                if q_cad and q_cad[0] == q_may[0] and \
                   abs(q_cad[1] - q_may[1]) > 0.10 * max(q_cad[1], q_may[1]):
                    sospechosos.append({
                        "mapeo": clave_mb, "sku": sku, "ean": ean,
                        "nombre_mayorista": nombre_may, "nombre_cadena": nom_cad,
                        "fuente_cadena": fte,
                    })
                    break

    # --- 5. Persistir ---
    try:
        # indent=2 para no reescribir el archivo entero: el auto-aprendizaje
        # viejo de main() usa indent=2 y ambos mecanismos escriben acá
        with open(MAPEO_BRUJULA_FILE, "w", encoding="utf-8") as f:
            json.dump(mb, f, ensure_ascii=False, indent=2)
        quality_dir = os.path.dirname(MATCHES_PENDIENTES_FILE)
        os.makedirs(quality_dir, exist_ok=True)
        with open(MATCHES_PENDIENTES_FILE, "w", encoding="utf-8") as f:
            json.dump({"fecha": datetime.now().isoformat(timespec="seconds"),
                       "pendientes": pendientes,
                       "rechazados": sorted(rechazados)}, f, ensure_ascii=False, indent=1)
        if sospechosos or aceptados:
            with open(MAPEOS_SOSPECHOSOS_FILE, "w", encoding="utf-8") as f:
                json.dump({"fecha": datetime.now().isoformat(timespec="seconds"),
                           "sospechosos": sospechosos,
                           "aceptados": sorted(aceptados)}, f, ensure_ascii=False, indent=1)
        elif os.path.isfile(MAPEOS_SOSPECHOSOS_FILE):
            os.remove(MAPEOS_SOSPECHOSOS_FILE)
    except OSError as e:
        print(f"  [WARN] No se pudo persistir expansion de mapeo: {e}")

    print(f"  Expansion via cadenas: +{n_auto['yaguar']} EANs Yaguar, "
          f"+{n_auto['maxiconsumo']} Maxiconsumo, +{n_auto['nini']} Nini aprendidos (persistidos), "
          f"{len(pendientes)} a revision, {len(sospechosos)} mapeos sospechosos")
    return n_auto, pendientes


def construir_catalogo(yaguar_data, maxicarre_data, maxiconsumo_data,
                       yag_sku_to_ean, mco_sku_to_ean,
                       ean_to_yag_sku, ean_to_mco_sku,
                       ean_to_master, nombre_norm_to_ean,
                       ean_to_familia=None,
                       nini_data=None, nini_sku_to_ean=None, ean_to_nini_sku=None):

    if ean_to_familia is None:
        ean_to_familia = {}
    nini_data = nini_data or []
    nini_sku_to_ean = nini_sku_to_ean if nini_sku_to_ean is not None else {}
    ean_to_nini_sku = ean_to_nini_sku if ean_to_nini_sku is not None else {}

    catalogo = {}   # prod_id -> entry

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def info_master(ean, fb_nombre, fb_sector):
        if ean and ean in ean_to_master:
            m = ean_to_master[ean]
            nombre = m["nombre"] if m["nombre"] and len(m["nombre"]) >= 3 else normalizar_nombre_display(fb_nombre)
            return nombre, m["sector"], m["categoria"], m["abc"]
        return normalizar_nombre_display(fb_nombre), normalizar_sector(fb_sector), "", ""

    def nuevo_producto(prod_id, ean, nombre, imagen, sector, subcategoria, abc=""):
        return {
            "id_unificado":   prod_id,
            "ean":            ean,
            "nombre_display": nombre,
            "imagen":         imagen,
            "sector":         sector,
            "subcategoria":   subcategoria,
            "abc":            abc,
            "precios":        {"yaguar": 0, "maxicarrefour": 0, "maxiconsumo": 0, "nini": 0,
                               "coto": 0, "carrefour": 0, "dia": 0,
                               "masonline": 0, "jumbo": 0},
            "fuentes":        {},
        }

    def resolver_ean(sku, sku_to_ean, nombre):
        """Obtiene EAN para un producto: CODIGOS primero, luego nombre->Maestro."""
        ean = sku_to_ean.get(str(sku).strip(), "")
        if not ean and nombre:
            ean = nombre_norm_to_ean.get(clave_nombre(nombre), "")
        return _ean_canonico(ean)

    # ------------------------------------------------------------------
    # PASO 1: Indexar Yaguar y Maxiconsumo por SKU y por nombre
    # ------------------------------------------------------------------
    yag_by_sku   = {}
    yag_by_clave = {}
    for p in yaguar_data:
        sku = str(p.get("sku", "")).strip()
        nom = p.get("nombre", "")
        if sku:
            yag_by_sku[sku] = p
        if nom:
            yag_by_clave[clave_nombre(nom)] = p

    mco_by_sku   = {}
    mco_by_clave = {}
    for p in maxiconsumo_data:
        if p.get("precio", 0) <= 0:
            continue
        sku = str(p.get("sku", "")).strip()
        nom = p.get("nombre", "")
        if sku:
            mco_by_sku[sku] = p
        if nom:
            mco_by_clave[clave_nombre(nom)] = p

    nini_by_sku   = {}
    nini_by_clave = {}
    for p in nini_data:
        sku = str(p.get("sku", "")).strip()
        nom = p.get("nombre", "")
        if p.get("precio", 0) <= 0:
            continue
        if sku:
            nini_by_sku[sku] = p
        if nom:
            nini_by_clave[clave_nombre(nom)] = p

    yag_merged  = set()   # SKUs de Yaguar ya procesados
    mco_merged  = set()   # SKUs de Maxiconsumo ya procesados
    nini_merged = set()   # SKUs de Nini ya procesados

    # ------------------------------------------------------------------
    # PASO 1b: Enriquecer ean_to_yag_sku y ean_to_mco_sku con Listado Maestro
    #   Tres estrategias en orden: campo ean externo → nombre exacto → Jaccard fuzzy.
    #   El índice fuzzy corre sobre nombre_norm_to_ean (25k+ entradas del Maestro)
    #   con threshold 0.60 — igual que enriquecer_eans.py pero sin depender del
    #   archivo preseleccionado por encontrar_mejor().
    # ------------------------------------------------------------------
    _FUZZ1B_TH   = 0.65
    _FUZZ1B_STOP = {"de", "la", "el", "en", "y", "x", "con", "por", "para",
                    "un", "una", "del", "los", "las", "al", "ml", "gr", "cc", "kg"}
    _fuzz_entries  = []   # (clave, word_set, ean)
    _fuzz_word_idx = defaultdict(list)
    for _clave, _ean in nombre_norm_to_ean.items():
        _ws = {w for w in _clave.split() if len(w) > 1 and w not in _FUZZ1B_STOP}
        if not _ws:
            continue
        _fi = len(_fuzz_entries)
        _fuzz_entries.append((_clave, _ws, _ean))
        for _fw in _ws:
            _fuzz_word_idx[_fw].append(_fi)

    _QTY_RE = re.compile(r"^\d+(?:gr|ml|kg|cc|un|ul)$")
    _PACK_UN_RE = re.compile(r"(\d+)un\b")

    def _pack_count(clave):
        m = _PACK_UN_RE.search(clave)
        return m.group(1) if m else ""

    def _familia_compat(id_a: str, id_b: str) -> bool:
        """False solo si ambos EANs tienen FAMILIA conocida y distinta."""
        fa = ean_to_familia.get(id_a, "")
        fb = ean_to_familia.get(id_b, "")
        if not fa or not fb:
            return True
        return fa == fb

    _CATEGORIA_TOKENS = {"gaseosa", "agua", "soda", "jugo", "cerveza", "vino", "aceite",
                         "arroz", "azucar", "harina", "leche", "yogur", "manteca", "queso",
                         "yerba", "cafe", "te", "sal", "fideos", "galletita", "galleta",
                         "chocolate", "caramelo", "golosina", "snack", "papel", "detergente",
                         "jabon", "shampoo", "desodorante", "lavandina", "suavizante", "servilleta",
                         "gin", "whisky", "fernet", "vodka", "ron", "sidra", "champagne",
                         "mayonesa", "ketchup", "mostaza", "salsa", "vinagre", "gelatina",
                         "polvo", "preparado", "mix", "pack", "combo", "duo"}

    def _primer_token_marca(clave_norm):
        """Primer token significativo que no sea categoria comun ni stop word."""
        for w in clave_norm.split():
            if len(w) > 2 and w not in _FUZZ1B_STOP and w not in _CATEGORIA_TOKENS:
                return w
        return ""

    def _fuzzy_ean_1b(nombre_prod):
        """Devuelve (ean, score). ean='' y score=0.0 si no supera el threshold."""
        _cl   = clave_nombre(nombre_prod)
        _ws_p = {w for w in _cl.split() if len(w) > 1 and w not in _FUZZ1B_STOP}
        if not _ws_p:
            return "", 0.0
        _qty_p = {w for w in _ws_p if _QTY_RE.match(w)}
        _marca_p = _primer_token_marca(_cl)
        _cands = set()
        for _w in _ws_p:
            for _i in _fuzz_word_idx.get(_w, []):
                _cands.add(_i)
        _best_sim = 0.0
        _best_ean = ""
        for _i in _cands:
            _clave_m, _ws_m, _ean = _fuzz_entries[_i]
            # Si ambos tienen tokens de cantidad (ej "500gr", "1500ml"), deben coincidir
            if _qty_p:
                _qty_m = {w for w in _ws_m if _QTY_RE.match(w)}
                if _qty_m and not (_qty_p & _qty_m):
                    continue
            # Pack count debe coincidir: "5un" vs "" → rechazar
            if _pack_count(_cl) != _pack_count(_clave_m):
                continue
            # Marca (primer token significativo) debe coincidir — evita MANAOS vs IVESS
            if _marca_p:
                _marca_m = _primer_token_marca(_clave_m)
                if _marca_m and _marca_p != _marca_m:
                    continue
            # Identificadores numéricos de variante (>=3 dígitos) deben coincidir — evita "1882" vs "Branca"
            _ids_num_p = {w for w in _ws_p if w.isdigit() and len(w) >= 3}
            if _ids_num_p:
                _ids_num_m = {w for w in _ws_m if w.isdigit() and len(w) >= 3}
                if _ids_num_m and not (_ids_num_p & _ids_num_m):
                    continue
            _inter = len(_ws_p & _ws_m)
            _union = len(_ws_p | _ws_m)
            _sim   = _inter / _union if _union else 0.0
            if _sim > _best_sim:
                _best_sim = _sim
                _best_ean = _ean
        return (_best_ean, _best_sim) if _best_sim >= _FUZZ1B_TH else ("", 0.0)

    _APRENDIZAJE_TH = 0.85
    _aprendizaje_yag = {}   # sku -> ean (matches fuzzy score >= 0.85)
    _aprendizaje_mco = {}   # sku -> ean

    ean_yag_nuevos = 0
    yag_sku_set = set(ean_to_yag_sku.values())
    for p in yaguar_data:
        sku = str(p.get("sku", "")).strip()
        if not sku or sku in yag_sku_set:
            continue
        ean_resuelto = str(p.get("ean", "") or "").strip()
        if not ean_resuelto or ean_resuelto in ("0", "None", "nan"):
            ean_resuelto = nombre_norm_to_ean.get(clave_nombre(p.get("nombre", "")), "")
            if ean_resuelto:
                _aprendizaje_yag[sku] = ean_resuelto  # lookup exacto -> guardar en mapeo
        if not ean_resuelto:
            ean_resuelto, _score = _fuzzy_ean_1b(p.get("nombre", ""))
            if ean_resuelto and _score >= _APRENDIZAJE_TH:
                _aprendizaje_yag[sku] = ean_resuelto
        ean_resuelto = _ean_canonico(ean_resuelto)
        if ean_resuelto and ean_resuelto not in ean_to_yag_sku:
            ean_to_yag_sku[ean_resuelto] = sku
            yag_sku_set.add(sku)
            ean_yag_nuevos += 1

    def _ean_brand_conflict(nombre_prod, ean):
        """Devuelve True si el EAN del scraper mapea a un nombre de marca diferente.
        Evita que un producto MOLTO quede asociado a un EAN de MATARAZZO, etc.
        """
        master_info = ean_to_master.get(ean, {})
        if not master_info:
            return False
        nombre_master = master_info.get("nombre", "")
        if not nombre_master:
            return False
        cl_p = clave_nombre(nombre_prod)
        cl_m = clave_nombre(nombre_master)
        ws_p = {w for w in cl_p.split() if len(w) > 2}
        ws_m = {w for w in cl_m.split() if len(w) > 2}
        if not ws_p or not ws_m:
            return False
        only_p = ws_p - ws_m
        only_m = ws_m - ws_p
        has_brand = lambda tokens: any(len(t) >= 4 and t.isalpha() for t in tokens)
        return has_brand(only_p) and has_brand(only_m)

    ean_mco_nuevos = 0
    ean_mco_brand_skip = 0
    mco_sku_set = set(ean_to_mco_sku.values())
    for p in maxiconsumo_data:
        sku = str(p.get("sku", "")).strip()
        if not sku or sku in mco_sku_set:
            continue
        ean_resuelto = str(p.get("ean", "") or "").strip()
        _from_fuzzy = False
        if not ean_resuelto or ean_resuelto in ("0", "None", "nan"):
            ean_resuelto = nombre_norm_to_ean.get(clave_nombre(p.get("nombre", "")), "")
            if ean_resuelto:
                _aprendizaje_mco[sku] = ean_resuelto  # lookup exacto -> guardar en mapeo
        elif _ean_brand_conflict(p.get("nombre", ""), ean_resuelto):
            ean_mco_brand_skip += 1
            ean_resuelto, _score = _fuzzy_ean_1b(p.get("nombre", ""))
            _from_fuzzy = True
        if not ean_resuelto:
            ean_resuelto, _score = _fuzzy_ean_1b(p.get("nombre", ""))
            _from_fuzzy = True
        if _from_fuzzy and ean_resuelto and _score >= _APRENDIZAJE_TH:
            _aprendizaje_mco[sku] = ean_resuelto
        ean_resuelto = _ean_canonico(ean_resuelto)
        if ean_resuelto and ean_resuelto not in ean_to_mco_sku:
            ean_to_mco_sku[ean_resuelto] = sku
            mco_sku_set.add(sku)
            ean_mco_nuevos += 1

    # Nini: mismo patrón que Yaguar (su API no entrega EAN crudo -- ver
    # investigación 29/07/2026 -- el conflicto de marca no aplica sin EAN propio).
    _aprendizaje_nini = {}
    ean_nini_nuevos = 0
    nini_sku_set = set(ean_to_nini_sku.values())
    for p in nini_data:
        sku = str(p.get("sku", "")).strip()
        if not sku or sku in nini_sku_set:
            continue
        ean_resuelto = nombre_norm_to_ean.get(clave_nombre(p.get("nombre", "")), "")
        if ean_resuelto:
            _aprendizaje_nini[sku] = ean_resuelto
        else:
            ean_resuelto, _score = _fuzzy_ean_1b(p.get("nombre", ""))
            if ean_resuelto and _score >= _APRENDIZAJE_TH:
                _aprendizaje_nini[sku] = ean_resuelto
        ean_resuelto = _ean_canonico(ean_resuelto)
        if ean_resuelto and ean_resuelto not in ean_to_nini_sku:
            ean_to_nini_sku[ean_resuelto] = sku
            nini_sku_set.add(sku)
            ean_nini_nuevos += 1

    print(f"  Paso 1b: +{ean_yag_nuevos} EANs Yaguar via Maestro, +{ean_mco_nuevos} EANs Maxiconsumo via Maestro "
          f"({ean_mco_brand_skip} descartados por conflicto de marca), +{ean_nini_nuevos} EANs Nini via Maestro")

    # Alias curados: renombrar claves de los indices ean->sku que vinieron de
    # CODIGOS.xlsx/mapeo con un EAN alias (ej. MCO Branca 750 mapeado al codigo
    # viejo 001193). Sin esto, el hub MCF busca por el canonico y no los ve.
    _alias_renombrados = 0
    if _ALIAS_EAN:
        for _idx in (ean_to_yag_sku, ean_to_mco_sku):
            for _e in list(_idx.keys()):
                _c = _ean_canonico(_e)
                if _c != _e:
                    _idx.setdefault(_c, _idx.pop(_e))
                    _alias_renombrados += 1
        if _alias_renombrados:
            print(f"  Alias EAN: {_alias_renombrados} claves de indices renombradas al canonico")

    # ------------------------------------------------------------------
    # PASO 2: MaxiCarrefour como HUB (100% EAN)
    #   Para cada producto MC busca en Yaguar y Maxiconsumo via CODIGOS
    # ------------------------------------------------------------------
    stats_mc = {"match_yag": 0, "match_mco": 0, "match_nini": 0, "nuevo": 0}

    for p in maxicarre_data:
        ean    = _ean_canonico(str(p.get("ean", "")).strip())
        nombre = p.get("nombre", "")
        precio = p.get("precio", 0)
        if not ean or not nombre:
            continue

        imagen_mc   = p.get("imagen", "")
        sector_raw  = p.get("sector", "")

        nombre_display, sector, subcategoria, abc = info_master(ean, nombre, sector_raw)

        entry = nuevo_producto(ean, ean, nombre_display, imagen_mc, sector, subcategoria, abc)
        entry["precios"]["maxicarrefour"] = precio
        entry["fuentes"]["maxicarrefour"] = {
            "nombre": nombre, "imagen": imagen_mc, "link": _carrefour_search_link(ean),
            "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", ""),
            # Folleto/Oportunidad Maxi: mismo criterio que Coto/Carrefour retail
            # (precio = efectivo, regular + texto para mostrar ambos)
            "precio_regular": p.get("precio_regular", 0),
            "oferta": p.get("oferta", ""),
        }

        # Buscar Yaguar via mapa inverso EAN->SKU
        yag_sku = ean_to_yag_sku.get(ean)
        if yag_sku and yag_sku in yag_by_sku:
            yag_p = yag_by_sku[yag_sku]
            yag_precio = yag_p.get("precio", 0)
            if yag_precio > 0:
                entry["precios"]["yaguar"] = yag_precio
            entry["fuentes"]["yaguar"] = {
                "nombre": yag_p.get("nombre", ""),
                "imagen": yag_p.get("imagen", ""),
                "sku":    yag_sku,
                "link":   yag_p.get("link", ""),
                "fecha_scraping": yag_p.get("fecha_scraping") or yag_p.get("fecha", ""),
            }
            yag_merged.add(yag_sku)
            stats_mc["match_yag"] += 1

        # Buscar Maxiconsumo via mapa inverso EAN->SKU
        mco_sku = ean_to_mco_sku.get(ean)
        if mco_sku and mco_sku in mco_by_sku:
            mco_p = mco_by_sku[mco_sku]
            mco_precio = mco_p.get("precio", 0)
            if mco_precio > 0:
                entry["precios"]["maxiconsumo"] = mco_precio
            entry["fuentes"]["maxiconsumo"] = {
                "nombre": mco_p.get("nombre", ""),
                "imagen": mco_p.get("imagen", ""),
                "sku":    mco_sku,
                "link":   mco_p.get("link", ""),
                "fecha_scraping": mco_p.get("fecha_scraping") or mco_p.get("fecha", ""),
            }
            mco_merged.add(mco_sku)
            stats_mc["match_mco"] += 1

        # Buscar Nini via mapa inverso EAN->SKU (aprendido, sin CODIGOS.xlsx propio)
        nini_sku = ean_to_nini_sku.get(ean)
        if nini_sku and nini_sku in nini_by_sku:
            nini_p = nini_by_sku[nini_sku]
            nini_precio = nini_p.get("precio", 0)
            if nini_precio > 0:
                entry["precios"]["nini"] = nini_precio
            entry["fuentes"]["nini"] = {
                "nombre": nini_p.get("nombre", ""),
                "sku":    nini_sku,
                "fecha_scraping": nini_p.get("fecha_scraping") or nini_p.get("fecha", ""),
            }
            nini_merged.add(nini_sku)
            stats_mc["match_nini"] += 1

        # Elegir la mejor imagen: Carrefour > Maxiconsumo > Yaguar (sin placeholders)
        # Nini no trae imagenes utilizables desde la API -- no entra a este pool.
        candidatas = [imagen_mc]
        if mco_sku and mco_sku in mco_by_sku:
            candidatas.append(mco_by_sku[mco_sku].get("imagen", ""))
        if yag_sku and yag_sku in yag_by_sku:
            candidatas.append(yag_by_sku[yag_sku].get("imagen", ""))
        entry["imagen"] = mejor_imagen(candidatas)

        # Si la imagen sigue siendo 0000- pero hay EAN, usar CDN Carrefour
        if "/0000-" in entry["imagen"] or not entry["imagen"]:
            entry["imagen"] = f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg"

        catalogo[ean] = entry
        stats_mc["nuevo"] += 1

    print(f"  MaxiCarrefour: {stats_mc['nuevo']} productos procesados")
    print(f"    -> Matches Yaguar via CODIGOS:      {stats_mc['match_yag']}")
    print(f"    -> Matches Maxiconsumo via CODIGOS: {stats_mc['match_mco']}")
    print(f"    -> Matches Nini via aprendizaje:    {stats_mc['match_nini']}")

    # ------------------------------------------------------------------
    # PASO 3: Yaguar - productos no mergeados con MaxiCarrefour
    # ------------------------------------------------------------------
    stats_yag = {"match_ean_catalogo": 0, "match_nombre_maestro": 0, "nuevo": 0}

    for p in yaguar_data:
        sku    = str(p.get("sku", "")).strip()
        nombre = p.get("nombre", "")
        precio = p.get("precio", 0)
        if not nombre or not sku:
            continue

        if sku in yag_merged:
            continue  # ya fue matcheado con MaxiCarrefour

        imagen    = p.get("imagen", "")
        sector_raw = mapear_sector_yaguar(p.get("categoria", ""))

        # Resolver EAN
        ean = resolver_ean(sku, yag_sku_to_ean, nombre)

        nombre_display, sector, subcategoria, abc = info_master(ean, nombre, sector_raw)

        if ean and ean in catalogo:
            # El EAN ya existe en catálogo (poco probable, pero por si acaso)
            catalogo[ean]["precios"]["yaguar"] = precio
            catalogo[ean]["fuentes"]["yaguar"] = {"nombre": nombre, "imagen": imagen, "sku": sku, "link": p.get("link", ""), "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", "")}
            if not catalogo[ean]["imagen"] or _es_placeholder(catalogo[ean]["imagen"]):
                catalogo[ean]["imagen"] = mejor_imagen([imagen, catalogo[ean]["imagen"]])
            if ean in nombre_norm_to_ean.values():
                stats_yag["match_ean_catalogo"] += 1
        else:
            prod_id = ean if ean else f"yaguar_{sku}"
            if prod_id not in catalogo:
                img_final = imagen
                if _es_placeholder(img_final) and ean:
                    img_final = f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg"
                entry = nuevo_producto(prod_id, ean, nombre_display, img_final, sector, subcategoria, abc)
                catalogo[prod_id] = entry
                stats_yag["nuevo"] += 1

            catalogo[prod_id]["precios"]["yaguar"] = precio
            catalogo[prod_id]["fuentes"]["yaguar"] = {"nombre": nombre, "imagen": imagen, "sku": sku, "link": p.get("link", ""), "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", "")}
            if ean:
                stats_yag["match_nombre_maestro"] += 1

        yag_merged.add(sku)

    print(f"  Yaguar (restantes): {stats_yag['nuevo']} nuevos, "
          f"{stats_yag['match_ean_catalogo']} match EAN, "
          f"{stats_yag['match_nombre_maestro']} con EAN via Maestro")

    # ------------------------------------------------------------------
    # PASO 3b: Nini - productos no mergeados con MaxiCarrefour
    #   Mismo tratamiento que Yaguar (PASO 3): la API de Nini no trae imagen
    #   ni link de producto, asi que esos campos quedan vacios.
    # ------------------------------------------------------------------
    stats_nini = {"match_ean_catalogo": 0, "match_nombre_maestro": 0, "nuevo": 0}

    for p in nini_data:
        sku    = str(p.get("sku", "")).strip()
        nombre = p.get("nombre", "")
        precio = p.get("precio", 0)
        if not nombre or not sku or precio <= 0:
            continue

        if sku in nini_merged:
            continue  # ya fue matcheado con MaxiCarrefour

        sector_raw = p.get("categoria", "")

        # Resolver EAN
        ean = resolver_ean(sku, nini_sku_to_ean, nombre)

        nombre_display, sector, subcategoria, abc = info_master(ean, nombre, sector_raw)

        if ean and ean in catalogo:
            catalogo[ean]["precios"]["nini"] = precio
            catalogo[ean]["fuentes"]["nini"] = {"nombre": nombre, "sku": sku, "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", "")}
            if ean in nombre_norm_to_ean.values():
                stats_nini["match_ean_catalogo"] += 1
        else:
            prod_id = ean if ean else f"nini_{sku}"
            if prod_id not in catalogo:
                img_final = f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg" if ean else ""
                entry = nuevo_producto(prod_id, ean, nombre_display, img_final, sector, subcategoria, abc)
                catalogo[prod_id] = entry
                stats_nini["nuevo"] += 1

            catalogo[prod_id]["precios"]["nini"] = precio
            catalogo[prod_id]["fuentes"]["nini"] = {"nombre": nombre, "sku": sku, "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", "")}
            if ean:
                stats_nini["match_nombre_maestro"] += 1

        nini_merged.add(sku)

    print(f"  Nini (restantes): {stats_nini['nuevo']} nuevos, "
          f"{stats_nini['match_ean_catalogo']} match EAN, "
          f"{stats_nini['match_nombre_maestro']} con EAN via Maestro")

    # ------------------------------------------------------------------
    # PASO 4: Maxiconsumo - productos no mergeados
    # ------------------------------------------------------------------
    stats_mco = {"match_ean_catalogo": 0, "match_nombre_yaguar": 0, "nuevo": 0}

    # Índice de claves de productos Yaguar sin EAN (para match por nombre)
    yag_clave_a_id = {}
    for prod_id, entry in catalogo.items():
        if not entry.get("ean") or prod_id.startswith("yaguar_"):
            clave = clave_nombre(entry["nombre_display"])
            if clave:
                yag_clave_a_id[clave] = prod_id

    for p in maxiconsumo_data:
        sku    = str(p.get("sku", "")).strip()
        nombre = p.get("nombre", "")
        precio = p.get("precio", 0)
        if not nombre or not sku or precio <= 0:
            continue

        if sku in mco_merged:
            continue  # ya matcheado con MaxiCarrefour

        imagen     = p.get("imagen", "")
        sector_raw = p.get("sector", "")

        # Resolver EAN
        ean = resolver_ean(sku, mco_sku_to_ean, nombre)

        nombre_display, sector, subcategoria, abc = info_master(ean, nombre, sector_raw)

        if ean and ean in catalogo:
            # EAN ya en catálogo
            catalogo[ean]["precios"]["maxiconsumo"] = precio
            catalogo[ean]["fuentes"]["maxiconsumo"] = {"nombre": nombre, "imagen": imagen, "sku": sku, "link": p.get("link", ""), "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", "")}
            if not catalogo[ean]["imagen"] or _es_placeholder(catalogo[ean]["imagen"]):
                catalogo[ean]["imagen"] = mejor_imagen([imagen, catalogo[ean]["imagen"]])
            stats_mco["match_ean_catalogo"] += 1
            mco_merged.add(sku)
            continue

        # Fallback: match por nombre con Yaguar sin EAN
        clave = clave_nombre(nombre_display)
        if clave in yag_clave_a_id:
            prod_id = yag_clave_a_id[clave]
            catalogo[prod_id]["precios"]["maxiconsumo"] = precio
            catalogo[prod_id]["fuentes"]["maxiconsumo"] = {"nombre": nombre, "imagen": imagen, "sku": sku, "link": p.get("link", ""), "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", "")}
            if not catalogo[prod_id]["imagen"]:
                catalogo[prod_id]["imagen"] = imagen
            stats_mco["match_nombre_yaguar"] += 1
            mco_merged.add(sku)
            continue

        # Producto nuevo
        prod_id = ean if ean else f"mco_{sku}"
        if prod_id not in catalogo:
            entry = nuevo_producto(prod_id, ean, nombre_display, imagen, sector, subcategoria, abc)
            catalogo[prod_id] = entry
            stats_mco["nuevo"] += 1

        catalogo[prod_id]["precios"]["maxiconsumo"] = precio
        catalogo[prod_id]["fuentes"]["maxiconsumo"] = {"nombre": nombre, "imagen": imagen, "sku": sku, "link": p.get("link", ""), "fecha_scraping": p.get("fecha_scraping") or p.get("fecha", "")}
        mco_merged.add(sku)

    print(f"  Maxiconsumo (restantes): {stats_mco['nuevo']} nuevos, "
          f"{stats_mco['match_ean_catalogo']} match EAN, "
          f"{stats_mco['match_nombre_yaguar']} match nombre Yaguar")

    # ------------------------------------------------------------------
    # PASO 5: Hunterprice bridge (triple Jaccard matching)
    #   Para cada producto de hunterprice:
    #     1. Buscar en MaxiCarrefour por nombre → obtener EAN → entry en catálogo
    #     2. Si le falta Yaguar: buscar en Yaguar scraper por nombre
    #     3. Si le falta Maxiconsumo: buscar en Maxiconsumo scraper por nombre
    #   Esto cubre productos que CODIGOS no pudo linkear por EAN.
    # ------------------------------------------------------------------
    hp_data = cargar_hunterprice()
    stats_hp = {"completados_yag": 0, "completados_mco": 0, "no_match_mc": 0}

    # Palabras de ruido para matching
    _STOP = {"x", "de", "la", "el", "y", "con", "sin", "pet", "pvc",
             "bot", "sdo", "fco", "brik", "p", "s", "en"}

    def _pals(clave):
        return {w for w in clave.split()
                if len(w) > 1 and w not in _STOP and not w.isdigit()}

    # Captura tanto números dentro de tokens de unidad ("1500ml", "500gr")
    # como números sueltos ("12", "4"). Dos grupos: uno con unidad, uno sin.
    _NUM_RE = re.compile(r"(\d+)(?:ml|gr|kg|un|cc)\b|\b(\d{2,5})\b")

    def _nums(clave):
        """Extrae números significativos (cantidad/tamaño) de una clave.
        Captura tanto '1500' de '1500ml' como números sueltos tipo '12'.
        """
        result = set()
        for m in _NUM_RE.finditer(clave):
            n = m.group(1) or m.group(2)
            if n:
                result.add(n)
        return result

    def _mejor_match(hp_ps, entries_list, word_index, threshold, hp_clave="", qty_ref=""):
        """
        Mejor match Jaccard en entries_list para hp_ps.
        Si el referente de cantidad (qty_ref > hp_clave cuando disponible) tiene números,
        el candidato debe compartir al menos uno. Esto evita cruzar tamaños distintos.
        qty_ref se usa para validar cantidad; si vacío, se cae a hp_clave.
        """
        cands = set()
        for _w in hp_ps:
            for _i in word_index.get(_w, []):
                cands.add(_i)
        if not cands:
            return None, 0.0
        # Preferir qty_ref (ej. nombre de MaxiCarrefour en catálogo) sobre hp_clave
        # para la validación de cantidad, ya que MC tiene el EAN correcto.
        _ref = qty_ref if qty_ref else hp_clave
        hp_numeros = _nums(_ref) if _ref else set()
        mejor_sim = 0.0
        mejor_val = None
        for _i in cands:
            _entry = entries_list[_i]
            _val   = _entry[0]
            _ps_c  = _entry[1]
            _clave_c = _entry[2] if len(_entry) > 2 else ""
            _inter = len(hp_ps & _ps_c)
            _union = len(hp_ps | _ps_c)
            _sim = _inter / _union if _union else 0.0
            if _sim < threshold:
                continue
            # Validar compatibilidad de cantidades
            # Si el referente tiene números y el candidato también, deben coincidir al menos 1
            if hp_numeros and _clave_c:
                _mc_numeros = _nums(_clave_c)
                if _mc_numeros and not (hp_numeros & _mc_numeros):
                    continue  # cantidades incompatibles
            # Pack count debe coincidir
            if _clave_c and _pack_count(_ref) != _pack_count(_clave_c):
                continue
            if _sim > mejor_sim:
                mejor_sim = _sim
                mejor_val = _val
        if not mejor_val:
            return None, mejor_sim
        return mejor_val, mejor_sim

    _TH = 0.50  # Jaccard mínimo

    # Índice invertido MaxiCarrefour: (ean, pals, clave)
    mc_entries_hp = []
    mc_word_idx   = defaultdict(list)
    for _p in maxicarre_data:
        _ean = str(_p.get("ean", "")).strip()
        _nom = _p.get("nombre", "")
        if not _ean or not _nom:
            continue
        _cl = clave_nombre(_nom)
        _ps = _pals(_cl)
        if not _ps:
            continue
        _i = len(mc_entries_hp)
        mc_entries_hp.append((_ean, _ps, _cl))
        for _w in _ps:
            mc_word_idx[_w].append(_i)

    # Índice invertido Yaguar: (producto, pals, clave)
    yag_entries_hp = []
    yag_word_idx   = defaultdict(list)
    for _p in yaguar_data:
        _nom = _p.get("nombre", "")
        if not _nom or _p.get("precio", 0) <= 0:
            continue
        _cl = clave_nombre(_nom)
        _ps = _pals(_cl)
        if not _ps:
            continue
        _i = len(yag_entries_hp)
        yag_entries_hp.append((_p, _ps, _cl))
        for _w in _ps:
            yag_word_idx[_w].append(_i)

    # Índice invertido Maxiconsumo: (producto, pals, clave)
    mco_entries_hp = []
    mco_word_idx   = defaultdict(list)
    for _p in maxiconsumo_data:
        _nom = _p.get("nombre", "")
        if not _nom or _p.get("precio", 0) <= 0:
            continue
        _cl = clave_nombre(_nom)
        _ps = _pals(_cl)
        if not _ps:
            continue
        _i = len(mco_entries_hp)
        mco_entries_hp.append((_p, _ps, _cl))
        for _w in _ps:
            mco_word_idx[_w].append(_i)

    for hp in hp_data:
        hp_nombre = (hp.get("Descripcion_Norm") or hp.get("Nombre_Unificado") or "").strip()
        if not hp_nombre:
            continue
        hp_ps = _pals(clave_nombre(hp_nombre))
        if not hp_ps:
            continue

        hp_tiene_yag = bool(hp.get("YAGUAR"))
        hp_tiene_mco = bool(hp.get("MAXICONSUMO"))
        if not hp_tiene_yag and not hp_tiene_mco:
            continue

        hp_clave = clave_nombre(hp_nombre)

        # PASO A: Encontrar EAN via MaxiCarrefour
        ean, sim_mc = _mejor_match(hp_ps, mc_entries_hp, mc_word_idx, _TH, hp_clave)
        if not ean:
            stats_hp["no_match_mc"] += 1
            continue

        if ean not in catalogo:
            continue
        entry = catalogo[ean]

        # Cantidad de referencia = nombre del producto en MaxiCarrefour (tiene EAN correcto).
        # Usarla en PASO B y C para no cruzar tamaños distintos cuando HP no tiene unidad.
        _mc_src_nombre = entry["fuentes"].get("maxicarrefour", {}).get("nombre", "")
        _qty_ref = clave_nombre(_mc_src_nombre) if _mc_src_nombre else hp_clave

        # PASO B: Completar Yaguar si falta
        if hp_tiene_yag and entry["precios"].get("yaguar", 0) == 0:
            # Primero via CODIGOS (ya intentado en PASO 2, pero por si acaso)
            yag_sku  = ean_to_yag_sku.get(ean)
            yag_prod = yag_by_sku.get(yag_sku) if yag_sku else None
            # Si no: Jaccard sobre Yaguar scraper (usar MC como referente de cantidad)
            if not yag_prod:
                yag_prod, _ = _mejor_match(hp_ps, yag_entries_hp, yag_word_idx, _TH, hp_clave, qty_ref=_qty_ref)
                yag_sku = str(yag_prod.get("sku", "")).strip() if yag_prod else ""
            if yag_prod and yag_prod.get("precio", 0) > 0:
                _yag_ean = yag_sku_to_ean.get(yag_sku, "")
                if not _familia_compat(ean, _yag_ean):
                    pass  # FAMILIA incompatible
                else:
                    entry["precios"]["yaguar"] = yag_prod["precio"]
                    entry["fuentes"]["yaguar"] = {
                        "nombre": yag_prod.get("nombre", ""),
                        "imagen": yag_prod.get("imagen", ""),
                        "sku":    yag_sku,
                        "link":   yag_prod.get("link", ""),
                        "fecha_scraping": yag_prod.get("fecha_scraping") or yag_prod.get("fecha", ""),
                    }
                    if yag_sku:
                        yag_merged.add(yag_sku)
                    stats_hp["completados_yag"] += 1

        # PASO C: Completar Maxiconsumo si falta
        if hp_tiene_mco and entry["precios"].get("maxiconsumo", 0) == 0:
            mco_sku  = ean_to_mco_sku.get(ean)
            mco_prod = mco_by_sku.get(mco_sku) if mco_sku else None
            if not mco_prod:
                mco_prod, _ = _mejor_match(hp_ps, mco_entries_hp, mco_word_idx, _TH, hp_clave, qty_ref=_qty_ref)
                mco_sku = str(mco_prod.get("sku", "")).strip() if mco_prod else ""
            if mco_prod and mco_prod.get("precio", 0) > 0:
                _mco_ean = mco_sku_to_ean.get(mco_sku, "")
                if not _familia_compat(ean, _mco_ean):
                    pass  # FAMILIA incompatible
                else:
                    entry["precios"]["maxiconsumo"] = mco_prod["precio"]
                    entry["fuentes"]["maxiconsumo"] = {
                        "nombre": mco_prod.get("nombre", ""),
                        "imagen": mco_prod.get("imagen", ""),
                        "sku":    mco_sku,
                        "link":   mco_prod.get("link", ""),
                        "fecha_scraping": mco_prod.get("fecha_scraping") or mco_prod.get("fecha", ""),
                    }
                    if mco_sku:
                        mco_merged.add(mco_sku)
                    stats_hp["completados_mco"] += 1

    print(f"  Hunterprice bridge: +{stats_hp['completados_yag']} Yaguar, "
          f"+{stats_hp['completados_mco']} Maxiconsumo | "
          f"{stats_hp['no_match_mc']} sin match MC")

    # ------------------------------------------------------------------
    # PASO 6: Post-proceso
    #   - Validación cruzada de precios (descarta outliers)
    #   - Eliminar productos sin precio
    #   - Reparar imágenes 0000- con CDN Carrefour (si tienen EAN)
    #   - Fusionar duplicados de nombre exacto
    # ------------------------------------------------------------------
    lista = list(catalogo.values())

    # ------ Filtro de precio mínimo absoluto ------
    # En Argentina 2026 ningún producto mayorista cuesta < $200.
    # Precios inferiores son artefactos de parsing (ej: "1.052701" en lugar de "1052.70").
    PRECIO_MINIMO = 200
    bajos_eliminados = 0
    for p in lista:
        for fuente in list(p["precios"]):
            if 0 < p["precios"][fuente] < PRECIO_MINIMO:
                p["precios"][fuente] = 0
                p["fuentes"].pop(fuente, None)
                bajos_eliminados += 1
    if bajos_eliminados:
        print(f"  Filtro mínimo (${PRECIO_MINIMO}): {bajos_eliminados} precios bajos eliminados")

    # ------ Filtro: unidades pequeñas con precios absurdos ------
    # Sachets <100ml o <50g con precio >$10,000 son errores de scraping (ej: Pantene 10ml a $190k)
    _SMALL_UNIT_RE = re.compile(r'\b([0-9]+)\s*(ml|cc|g|gr)\b', re.IGNORECASE)
    _PRECIO_UNITARIO_MAX = 10_000
    techo_unitario = 0
    for p in lista:
        m = _SMALL_UNIT_RE.search(p.get("nombre_display", ""))
        if not m:
            continue
        cant, unidad = int(m.group(1)), m.group(2).lower()
        if (unidad in ('ml', 'cc') and cant < 100) or (unidad in ('g', 'gr') and cant < 50):
            for fuente in list(p["precios"].keys()):
                if p["precios"][fuente] > _PRECIO_UNITARIO_MAX:
                    p["precios"][fuente] = 0
                    p["fuentes"].pop(fuente, None)
                    techo_unitario += 1
    if techo_unitario:
        print(f"  Filtro tamaño pequeño (<100ml/<50g >$10k): {techo_unitario} precios absurdos eliminados")

    # ------ Validación cruzada de precios ------
    # Si un producto tiene múltiples fuentes y un precio es >5x más barato
    # que el resto, se descarta como outlier (scraping error).
    # Solo aplica cuando hay al menos 2 fuentes para comparar.
    precios_descartados = 0
    sospechosos_marcados = 0
    for p in lista:
        precios_activos = {k: v for k, v in p["precios"].items() if v > 0}
        if len(precios_activos) < 2:
            continue
        vals = list(precios_activos.values())
        mediana = sorted(vals)[len(vals) // 2]
        min_val = min(vals)
        for fuente, precio in list(precios_activos.items()):
            # Outlier hacia abajo: solo con 3+ precios para evitar descartar el precio
            # correcto cuando el outlier alto (ej. Yaguar display) distorsiona la mediana.
            if len(vals) >= 3 and precio < mediana / 4 and mediana > 800:
                p["precios"][fuente] = 0
                if fuente in p["fuentes"]:
                    del p["fuentes"][fuente]
                precios_descartados += 1
            # Outlier hacia arriba con 3+ fuentes: la mediana es robusta (2 fuentes
            # coinciden), así que un precio > 2.5x la mediana es error de scraping o
            # producto sin stock sobrevaluado (caso Queso La Paulina: MC $4.809 vs
            # ~$1.550 en Yaguar/Carrefour, 3.1x). El umbral 2.5x ignora la dif. IIBB
            # del ~3% (regla 09) que nunca llega ni a 1.1x.
            elif len(vals) >= 3 and precio > mediana * 2.5:
                p["precios"][fuente] = 0
                if fuente in p["fuentes"]:
                    del p["fuentes"][fuente]
                precios_descartados += 1
            # Outlier hacia arriba extremo (packs/displays 10x-1300x): descartar siempre,
            # incluso con solo 2 fuentes.
            elif precio > min_val * 10 and min_val > 0:
                p["precios"][fuente] = 0
                if fuente in p["fuentes"]:
                    del p["fuentes"][fuente]
                precios_descartados += 1
        # Caso ambiguo de 2 fuentes: con solo dos precios no se puede saber cuál miente,
        # así que NO se descarta. Se marca la fuente más cara como sospechosa cuando
        # supera 2.5x a la otra, para revisión manual y para que el frontend pueda avisar.
        activos_post = {k: v for k, v in p["precios"].items() if v > 0}
        if len(activos_post) == 2:
            v_ord = sorted(activos_post.values())
            if v_ord[1] > v_ord[0] * 2.5:
                cara = max(activos_post, key=activos_post.get)
                if cara in p["fuentes"]:
                    p["fuentes"][cara]["precio_sospechoso"] = True
                    sospechosos_marcados += 1
    if precios_descartados:
        print(f"  Validación cruzada: {precios_descartados} precios outlier descartados")
    if sospechosos_marcados:
        print(f"  Precios sospechosos (2 fuentes, ratio >2.5x): {sospechosos_marcados} marcados para revisión")

    # Eliminar sin precio
    lista = [p for p in lista if any(v > 0 for v in p["precios"].values())]

    # Reparar imágenes placeholder
    for p in lista:
        img = p.get("imagen", "")
        ean = p.get("ean", "")
        if _es_placeholder(img) and ean:
            p["imagen"] = f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg"

    def _fusionar_grupo(items):
        """Fusiona una lista de productos al mejor representante.

        Base: EAN con respaldo en el Maestro > EAN a secas > sin EAN. Caso leche
        14/07/2026: un SKU Yaguar discontinuado con EAN muerto de CODIGOS.xlsx
        (sin respaldo en ningun lado) le ganaba como base al SKU vigente, y el
        producto quedaba aislado del merge de cadenas.
        Colision de fuente (ambos registros tienen la misma fuente): gana el de
        fecha_scraping mas reciente — antes se descartaba silenciosamente el del
        absorbido y podia publicarse el precio de un SKU discontinuado.
        """
        def _rango_base(x):
            e = x.get("ean", "")
            return 2 if (e and e in ean_to_master) else (1 if e else 0)
        base = max(items, key=_rango_base)
        for item in items:
            if item is base:
                continue
            for fuente, precio in item["precios"].items():
                if precio <= 0:
                    continue
                info_item = item.get("fuentes", {}).get(fuente)
                if base["precios"].get(fuente, 0) == 0:
                    base["precios"][fuente] = precio
                    if info_item and fuente not in base["fuentes"]:
                        base["fuentes"][fuente] = info_item
                elif info_item:
                    # fechas ISO (YYYY-MM-DD): comparacion lexicografica valida
                    f_base = str(base.get("fuentes", {}).get(fuente, {}).get("fecha_scraping", ""))
                    f_item = str(info_item.get("fecha_scraping", ""))
                    if f_item > f_base:
                        base["precios"][fuente] = precio
                        base["fuentes"][fuente] = info_item
            for fuente, info in item.get("fuentes", {}).items():
                if fuente not in base["fuentes"]:
                    base["fuentes"][fuente] = info
            if _es_placeholder(base.get("imagen", "")):
                if item.get("imagen") and not _es_placeholder(item.get("imagen", "")):
                    base["imagen"] = item["imagen"]
            if not base.get("abc") and item.get("abc"):
                base["abc"] = item["abc"]
        return base

    def _es_sintetico(prod_id):
        return str(prod_id).startswith(("yaguar_", "mco_", "nini_"))

    # Paso 6a: Fusionar duplicados de nombre_display exacto
    por_nombre = defaultdict(list)
    for p in lista:
        por_nombre[p["nombre_display"]].append(p)

    lista_paso6a = []
    for nombre, items in por_nombre.items():
        if len(items) == 1:
            lista_paso6a.append(items[0])
        else:
            lista_paso6a.append(_fusionar_grupo(items))

    # Paso 6b: Fusionar duplicados por nombre normalizado
    #   Solo fusiona si al menos uno tiene ID sintético (yaguar_XXX / mco_XXX),
    #   lo que indica que una fuente no pudo resolver el EAN del mismo producto.
    #   Si ambos tienen EAN real, son SKUs genuinamente distintos → no tocar.
    por_clave = defaultdict(list)
    for p in lista_paso6a:
        por_clave[clave_nombre(p["nombre_display"])].append(p)

    lista_final = []
    fusiones_norm = 0
    for clave_norm, items in por_clave.items():
        if len(items) == 1:
            lista_final.append(items[0])
            continue
        hay_sintetico = any(_es_sintetico(p["id_unificado"]) for p in items)
        if not hay_sintetico:
            # Todos tienen EAN real → SKUs genuinamente distintos, mantener separados
            lista_final.extend(items)
        else:
            ids = [p["id_unificado"] for p in items]
            familia_ok = all(
                _familia_compat(ids[i], ids[j])
                for i in range(len(ids)) for j in range(i + 1, len(ids))
            )
            if not familia_ok:
                lista_final.extend(items)
            else:
                lista_final.append(_fusionar_grupo(items))
                fusiones_norm += 1

    if fusiones_norm:
        print(f"  Paso 6b: {fusiones_norm} duplicados por nombre normalizado fusionados")

    # ------------------------------------------------------------------
    # PASO 6c: Fusión fuzzy de productos complementarios
    #   Para cada producto con precios faltantes, busca en los productos
    #   que tienen esa(s) fuente(s) faltante(s) usando Jaccard > 0.82.
    #   Prioridad de base: maxicarrefour (tiene EAN) > yaguar > maxiconsumo.
    # ------------------------------------------------------------------
    _STOP6 = {"de", "la", "el", "y", "con", "sin", "pet", "pvc",
              "bot", "sdo", "fco", "brik", "p", "s", "en"}
    # Mismo patrón que _NUM_RE: captura números dentro de unidades y sueltos
    _NUM6  = re.compile(r"(\d+)(?:ml|gr|kg|un|cc)\b|\b(\d{2,6})\b")

    def _w6(clave):
        return {w for w in clave.split() if len(w) > 1 and w not in _STOP6 and not w.isdigit()}

    def _n6(clave):
        result = set()
        for m in _NUM6.finditer(clave):
            n = m.group(1) or m.group(2)
            if n:
                result.add(n)
        return result

    _TH6 = 0.75  # subido de 0.65 — evita fusionar "Fernet 1882" con "Fernet Branca"

    def _buscar_candidato(ws_p, ns_p, index_entries, index_wi, usados, cl_p="", id_p="",
                          ean_p="", precios_p=None):
        """Devuelve (idx_en_lista_final, sim) del mejor match fuzzy."""
        cands = set()
        for w in ws_p:
            for ei in index_wi.get(w, []):
                cands.add(ei)
        mejor_sim = 0.0
        mejor_idx = None
        for ei in cands:
            lf_idx, ws_c, cl_c, ns_c = index_entries[ei]
            if lf_idx in usados:
                continue
            inter = len(ws_p & ws_c)
            union = len(ws_p | ws_c)
            sim   = inter / union if union else 0.0
            if sim < _TH6:
                continue
            if ns_p and ns_c and not (ns_p & ns_c):
                continue  # cantidades incompatibles
            if _pack_count(cl_p) != _pack_count(cl_c):
                continue  # pack count incompatible
            cand = lista_final[lf_idx]
            id_c = cand["id_unificado"]
            # Guard EAN (mismo criterio que 6b): dos EANs reales distintos = productos
            # genuinamente distintos, aunque el nombre dé Jaccard alto. Caso Playadito
            # 14/07/2026: "X500 g" (911) vs "500 g LATA" (089) daban sim=0.75 y se
            # fusionaban, borrando la variante comun del catalogo.
            ean_c = cand.get("ean", "")
            if ean_p and ean_c and ean_p != ean_c:
                _stats6c["ean_distinto"] += 1
                continue
            # Complementariedad estricta: si ambos ya tienen precio en la MISMA fuente
            # (SKUs distintos de esa fuente), no son el mismo producto — fusionar
            # descartaria uno de los dos registros en silencio.
            if precios_p and any(v > 0 and cand["precios"].get(f, 0) > 0
                                 for f, v in precios_p.items()):
                _stats6c["fuente_solapada"] += 1
                continue
            if not _familia_compat(id_p, id_c):
                continue  # FAMILIAs distintas → presentaciones incompatibles
            # Identificadores numéricos de variante (>=3 dígitos) deben coincidir — evita "1882" vs "Branca"
            _ids_num_p = {w for w in cl_p.split() if w.isdigit() and len(w) >= 3}
            if _ids_num_p:
                _ids_num_c = {w for w in cl_c.split() if w.isdigit() and len(w) >= 3}
                if _ids_num_c and not (_ids_num_p & _ids_num_c):
                    continue
            if sim > mejor_sim:
                mejor_sim = sim
                mejor_idx = lf_idx
        return mejor_idx, mejor_sim

    # Construir índices por fuente
    def _build_index(fuente, lista):
        """(idx_lista_final, ws, cl, ns) para productos que TIENEN fuente y solo fuente."""
        entries = []
        wi      = defaultdict(list)
        for idx, p in enumerate(lista):
            if p["precios"].get(fuente, 0) <= 0:
                continue
            cl = clave_nombre(p["nombre_display"])
            ws = _w6(cl)
            ns = _n6(cl)
            if not ws:
                continue
            ei = len(entries)
            entries.append((idx, ws, cl, ns))
            for w in ws:
                wi[w].append(ei)
        return entries, wi

    mc_idx_entries,   mc_idx_wi   = _build_index("maxicarrefour", lista_final)
    yag_idx_entries,  yag_idx_wi  = _build_index("yaguar", lista_final)
    mco_idx_entries,  mco_idx_wi  = _build_index("maxiconsumo", lista_final)
    nini_idx_entries, nini_idx_wi = _build_index("nini", lista_final)

    fusiones_fuzzy = 0
    usados_como_base = set()   # índices de lista_final que ya absorbieron algo
    _stats6c = {"ean_distinto": 0, "fuente_solapada": 0}

    # parche: { idx_eliminado: idx_base }
    parches = {}

    for idx_p, p in enumerate(lista_final):
        if idx_p in usados_como_base or idx_p in parches:
            continue
        pr = p["precios"]
        tiene_mc   = pr.get("maxicarrefour", 0) > 0
        tiene_yag  = pr.get("yaguar", 0) > 0
        tiene_mco  = pr.get("maxiconsumo", 0) > 0
        tiene_nini = pr.get("nini", 0) > 0
        n_fuentes = sum([tiene_mc, tiene_yag, tiene_mco, tiene_nini])
        if n_fuentes == 4:
            continue  # completo

        cl_p = clave_nombre(p["nombre_display"])
        ws_p = _w6(cl_p)
        ns_p = _n6(cl_p)
        if not ws_p:
            continue

        # Buscar fuentes faltantes en orden de prioridad
        for fuente_falt, entries_f, wi_f in [
            ("maxicarrefour", mc_idx_entries,   mc_idx_wi),
            ("yaguar",        yag_idx_entries,  yag_idx_wi),
            ("maxiconsumo",   mco_idx_entries,  mco_idx_wi),
            ("nini",          nini_idx_entries, nini_idx_wi),
        ]:
            if pr.get(fuente_falt, 0) > 0:
                continue  # ya tiene esta fuente

            lf_idx, sim = _buscar_candidato(ws_p, ns_p, entries_f, wi_f, usados_como_base | set(parches.keys()) | {idx_p}, cl_p=cl_p, id_p=p["id_unificado"], ean_p=p.get("ean", ""), precios_p=pr)
            if lf_idx is None:
                continue

            p_cand = lista_final[lf_idx]
            # Solo absorber si el candidato tiene SOLO esa fuente (o pocas fuentes)
            # → evitar partir un producto ya bien matcheado
            n_cand = sum(1 for v in p_cand["precios"].values() if v > 0)
            if n_cand > 2:
                continue  # candidato ya tiene muchos precios, no arriesgar

            # ¿Cuál es la "base" (el que tiene EAN)?
            p_tiene_ean = bool(p.get("ean"))
            cand_tiene_ean = bool(p_cand.get("ean"))
            if cand_tiene_ean and not p_tiene_ean:
                # Cand es la base, p se funde en cand
                parches[idx_p] = lf_idx
                usados_como_base.add(lf_idx)
            elif p_tiene_ean:
                # p es la base, cand se funde en p
                parches[lf_idx] = idx_p
                usados_como_base.add(idx_p)
            else:
                # Ninguno tiene EAN: base = el que tiene más fuentes
                if n_cand >= n_fuentes:
                    parches[idx_p] = lf_idx
                    usados_como_base.add(lf_idx)
                else:
                    parches[lf_idx] = idx_p
                    usados_como_base.add(idx_p)
            fusiones_fuzzy += 1
            break  # una fusión por producto es suficiente

    # Aplicar parches
    _auditoria_6c = []
    for idx_elim, idx_base in parches.items():
        p_elim = lista_final[idx_elim]
        p_base = lista_final[idx_base]
        for fuente, precio in p_elim["precios"].items():
            if precio > 0 and p_base["precios"].get(fuente, 0) == 0:
                p_base["precios"][fuente] = precio
        for fuente, info in p_elim.get("fuentes", {}).items():
            if fuente not in p_base["fuentes"]:
                p_base["fuentes"][fuente] = info.copy()
        if not p_base.get("abc") and p_elim.get("abc"):
            p_base["abc"] = p_elim["abc"]
        p_elim["_eliminar"] = True
        _auditoria_6c.append({
            "base_id":     p_base["id_unificado"],
            "base_nombre": p_base["nombre_display"],
            "absorbido_id":     p_elim["id_unificado"],
            "absorbido_nombre": p_elim["nombre_display"],
            "fuentes_movidas":  sorted(p_elim.get("fuentes", {}).keys()),
        })

    lista_final = [p for p in lista_final if not p.get("_eliminar")]

    # Auditoria de fusiones 6c (se pisa en cada corrida) — revisable por auditor-catalogo
    try:
        quality_dir = os.path.join(BASE_DIR, "data", "quality")
        os.makedirs(quality_dir, exist_ok=True)
        with open(os.path.join(quality_dir, "fusiones_6c.json"), "w", encoding="utf-8") as f:
            json.dump({"fecha": datetime.now().isoformat(timespec="seconds"),
                       "aplicadas": len(_auditoria_6c),
                       "descartadas_ean_distinto": _stats6c["ean_distinto"],
                       "descartadas_fuente_solapada": _stats6c["fuente_solapada"],
                       "fusiones": _auditoria_6c}, f, ensure_ascii=False, indent=1)
    except OSError as e:
        print(f"  [WARN] No se pudo escribir fusiones_6c.json: {e}")

    print(f"  Paso 6c: {fusiones_fuzzy} fusiones fuzzy complementarias "
          f"({_stats6c['ean_distinto']} candidatos descartados por EAN distinto, "
          f"{_stats6c['fuente_solapada']} por fuente solapada)")

    # ------------------------------------------------------------------
    # PASO 6d: Validación de cantidad entre fuentes (cleanup defensivo)
    #   Para cada producto con 2+ fuentes, extrae la cantidad CANÓNICA
    #   (volumen->ml, peso->g) del nombre de cada fuente. Si una fuente
    #   difiere >10% del "ancla" (MC > Yaguar > MCO) en la misma dimensión,
    #   son tamaños distintos: el link de esa fuente lleva a OTRO producto
    #   (ej. Glaciar 1.5L vs 2L) -> se elimina. El 10% tolera ruido de
    #   parsing (950ml vs 930ml = 2%) sin dejar pasar variantes reales.
    #
    #   Sub-filtro de pack NxM: si una fuente usa formato "N x M unidad"
    #   (ej. "32UNX13GR", "24X10 GR") y el ancla no, son unidades distintas
    #   (pack vs unidad individual) -> eliminar esa fuente directamente.
    # ------------------------------------------------------------------
    # Patron de pack multiplo: N (UN opcionalmente) x M unidad
    # Ejemplos: '32UNX13GR', '24X10 GR', '6X1L', '12X500ML'
    _PACK_NxM_RE = re.compile(
        r"(\d+)\s*(?:[uU][nN])?\s*[xX]\s*(\d+)\s*"
        r"(?:[gG][rR]?[sS]?|[mM][lL]|[cC][cC]|[kK][gG]|[lL][tT]?[sS]?)",
        re.IGNORECASE
    )

    def _tiene_pack_nxm(nombre_crudo):
        """True si el nombre contiene formato pack multiplo (NxM con unidad)."""
        return bool(_PACK_NxM_RE.search(nombre_crudo or ""))

    fuentes_eliminadas_6d = 0
    _ANCHOR_ORDER = ["maxicarrefour", "yaguar", "maxiconsumo", "nini"]

    for p in lista_final:
        precios_activos = {k: v for k, v in p["precios"].items() if v > 0}
        if len(precios_activos) < 2:
            continue
        fuentes = p.get("fuentes", {})

        # Elegir el ancla: la fuente con mayor confianza (MC > Yaguar > MCO)
        ancla = None
        for _f in _ANCHOR_ORDER:
            if precios_activos.get(_f, 0) > 0:
                ancla = _f
                break
        if not ancla:
            continue

        ancla_nombre = fuentes.get(ancla, {}).get("nombre", "")
        if not ancla_nombre:
            continue

        ancla_es_pack = _tiene_pack_nxm(ancla_nombre)

        for fuente in list(precios_activos.keys()):
            if fuente == ancla:
                continue
            src_nombre = fuentes.get(fuente, {}).get("nombre", "")
            if not src_nombre:
                continue

            # Sub-filtro pack NxM: si una fuente es pack y la otra no -> mismatch
            src_es_pack = _tiene_pack_nxm(src_nombre)
            if src_es_pack != ancla_es_pack:
                p["precios"][fuente] = 0
                if fuente in p["fuentes"]:
                    del p["fuentes"][fuente]
                fuentes_eliminadas_6d += 1
                continue

        # Re-leer precios activos tras el filtro de pack (puede haber cambiado)
        precios_activos = {k: v for k, v in p["precios"].items() if v > 0}
        if len(precios_activos) < 2:
            continue

        anc_cant = _cantidad_canonica(ancla_nombre)
        if not anc_cant:
            continue  # ancla sin info de cantidad: no validar

        for fuente in list(precios_activos.keys()):
            if fuente == ancla:
                continue
            src_nombre = fuentes.get(fuente, {}).get("nombre", "")
            if not src_nombre:
                continue
            src_cant = _cantidad_canonica(src_nombre)
            if not src_cant:
                continue  # fuente sin cantidad: no sancionar
            # Misma dimensión (vol vs vol, peso vs peso) y diferencia >10% ->
            # son tamaños distintos: el link de esa fuente va a otro producto.
            if src_cant[0] == anc_cant[0]:
                mayor = max(src_cant[1], anc_cant[1])
                menor = max(1, min(src_cant[1], anc_cant[1]))
                if mayor / menor > 1.10:
                    p["precios"][fuente] = 0
                    if fuente in p["fuentes"]:
                        del p["fuentes"][fuente]
                    fuentes_eliminadas_6d += 1

    if fuentes_eliminadas_6d:
        print(f"  Paso 6d: {fuentes_eliminadas_6d} fuentes con cantidad incompatible eliminadas")

    # Eliminar productos que quedaron sin precio tras la limpieza 6d
    lista_final = [p for p in lista_final if any(v > 0 for v in p["precios"].values())]

    # ------------------------------------------------------------------
    # PASO 6e: Segunda validación cruzada post-fusiones
    #   Paso 6c puede crear pares (precio_correcto, precio_malformado).
    #   La primera validación (antes de 6c) no los ve. Esta segunda pasada
    #   los elimina con el mismo criterio: outlier < mediana/4 o > mediana*10.
    # ------------------------------------------------------------------
    post_descartados = 0
    for p in lista_final:
        precios_activos = {k: v for k, v in p["precios"].items() if v > 0}
        if len(precios_activos) < 2:
            continue
        vals = list(precios_activos.values())
        mediana = sorted(vals)[len(vals) // 2]
        min_val = min(vals)
        for fuente, precio in list(precios_activos.items()):
            if len(vals) >= 3 and precio < mediana / 2.5 and mediana > 800:
                p["precios"][fuente] = 0
                p["fuentes"].pop(fuente, None)
                post_descartados += 1
            elif precio > min_val * 10 and min_val > 0:
                p["precios"][fuente] = 0
                p["fuentes"].pop(fuente, None)
                post_descartados += 1
    if post_descartados:
        print(f"  Paso 6e: {post_descartados} outliers post-fusión descartados")

    lista_final = [p for p in lista_final if any(v > 0 for v in p["precios"].values())]

    # ------------------------------------------------------------------
    # PASO 6f: Outlier Maxiconsumo (.claude/rules/08-precios-sin-stock.md)
    #   MC muestra precio aunque el producto este sin stock, y a veces el
    #   scraper captura precio de bulto. Auditoria 11/06: 23 casos donde
    #   MC era 2.5x-11x la mediana de las otras fuentes.
    # ------------------------------------------------------------------
    OUTLIER_MC_RATIO = 2.5
    mc_descartados = 0
    for p in lista_final:
        precio_mc = p["precios"].get("maxiconsumo", 0)
        if precio_mc <= 0:
            continue
        otras = [v for k, v in p["precios"].items() if k != "maxiconsumo" and v > 0]
        if not otras:
            continue
        mediana_otras = sorted(otras)[len(otras) // 2]
        if mediana_otras > 0 and precio_mc > mediana_otras * OUTLIER_MC_RATIO:
            print(f"  Outlier MC descartado: {p.get('nombre_display', '')[:55]} "
                  f"MC=${precio_mc:,.0f} vs mediana otras=${mediana_otras:,.0f}")
            p["precios"]["maxiconsumo"] = 0
            p["fuentes"].pop("maxiconsumo", None)
            mc_descartados += 1
    print(f"  PRECIOS SOSPECHOSOS MC descartados: {mc_descartados}")

    lista_final = [p for p in lista_final if any(v > 0 for v in p["precios"].values())]

    # ------------------------------------------------------------------
    # PASO 6g: Flag de comparacion sospechosa (ahorro > 60% entre fuentes)
    #   Casi siempre es un match incorrecto o un precio mal capturado, pero
    #   con 2 fuentes no se sabe cual esta mal -> no descartar, flagear.
    #   El frontend excluye estos productos de Top Bombas (lib/data.ts).
    # ------------------------------------------------------------------
    sospechosos = 0
    for p in lista_final:
        vals = [v for v in p["precios"].values() if v > 0]
        if len(vals) >= 2 and min(vals) < max(vals) * 0.4:
            p["precio_sospechoso"] = True
            sospechosos += 1
    if sospechosos:
        print(f"  Paso 6g: {sospechosos} productos con ahorro >60% flageados (excluidos de bombas)")

    for p in lista_final:
        if "familia" not in p:
            ean = p.get("id_unificado", "")
            p["familia"] = ean_to_master.get(ean, {}).get("familia", "")

    return lista_final, _aprendizaje_yag, _aprendizaje_mco, _aprendizaje_nini


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("ACTUALIZADOR DE CATALOGO v3.0 - Brujula de Precios")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    print("\nCargando tablas de referencia (Excel)...")
    (yag_sku_to_ean, mco_sku_to_ean, nini_sku_to_ean,
     ean_to_yag_sku, ean_to_mco_sku, ean_to_nini_sku,
     ean_to_master, nombre_norm_to_ean,
     ean_to_familia) = cargar_excel_referencia()

    print("\nCargando datos de scrapers (mejor archivo por cantidad)...")
    yaguar      = cargar_yaguar()
    maxicarre   = cargar_maxicarrefour()
    maxiconsumo = cargar_maxiconsumo()
    nini        = cargar_nini()
    coto        = cargar_cadena("coto", "Coto")
    carrefour   = cargar_cadena("carrefour", "Carrefour")
    dia         = cargar_cadena("dia", "Dia")
    masonline   = cargar_cadena("masonline", "Masonline")
    jumbo       = cargar_cadena("jumbo", "Jumbo")

    print("\nExpandiendo mapeo SKU->EAN con nombres frescos de cadenas...")
    expandir_mapeo_con_cadenas(
        yaguar, maxiconsumo,
        [("coto", coto), ("carrefour", carrefour), ("dia", dia),
         ("masonline", masonline), ("jumbo", jumbo),
         ("maxicarrefour", maxicarre)],
        yag_sku_to_ean, mco_sku_to_ean,
        ean_to_yag_sku, ean_to_mco_sku,
        nombre_norm_to_ean, ean_to_master,
        nini_data=nini, nini_sku_to_ean=nini_sku_to_ean, ean_to_nini_sku=ean_to_nini_sku,
    )

    print("\nConstruyendo catálogo unificado...")
    catalogo, aprendizaje_yag, aprendizaje_mco, aprendizaje_nini = construir_catalogo(
        yaguar, maxicarre, maxiconsumo,
        yag_sku_to_ean, mco_sku_to_ean,
        ean_to_yag_sku, ean_to_mco_sku,
        ean_to_master, nombre_norm_to_ean,
        ean_to_familia,
        nini_data=nini, nini_sku_to_ean=nini_sku_to_ean, ean_to_nini_sku=ean_to_nini_sku,
    )

    # ------ ABC fallback algorítmico ------
    # Productos sin ABC del Maestro → clasificar por precio promedio (percentiles 30/40/30).
    # abc_fuente="calculado" distingue estos del Maestro para análisis posterior.
    sin_abc = [p for p in catalogo if not p.get("abc")]
    if sin_abc:
        precios_prom = []
        for p in sin_abc:
            vals = [v for v in p["precios"].values() if v > 0]
            if vals:
                precios_prom.append((p, sum(vals) / len(vals)))
        precios_prom.sort(key=lambda x: x[1])
        n = len(precios_prom)
        for i, (p, _) in enumerate(precios_prom):
            if i >= n * 0.7:
                p["abc"] = "A"
            elif i >= n * 0.3:
                p["abc"] = "B"
            else:
                p["abc"] = "C"
            p["abc_fuente"] = "calculado"
        abc_calc_total = len(precios_prom)
        print(f"  ABC fallback calculado:       {abc_calc_total} productos sin clasificacion Maestro")

    # ------ Cadenas minoristas: Coto + Carrefour retail (100% EAN) ------
    # Se agregan DESPUES de construir_catalogo a proposito: las validaciones
    # internas (outliers 6e/6f, sospechosos 6g, cantidad 6d) comparan precios
    # entre si asumiendo mayoristas; las cadenas son legitimamente mas caras
    # (gondola) y adentro generarian falsos positivos. Fuera del constructor no
    # contaminan ni pueden ser descartadas. Decision de producto: solo agregan
    # precio a EANs que ya existen — los exclusivos quedan para la Fase B.
    cadenas = [("coto", "Coto", coto), ("carrefour", "Carrefour", carrefour), ("dia", "Dia", dia),
               ("masonline", "Masonline", masonline), ("jumbo", "Jumbo", jumbo)]
    if any(data for _, _, data in cadenas):
        idx_ean = {}
        for p in catalogo:
            ean_p = str(p.get("ean", "")).strip()
            if ean_p:
                idx_ean.setdefault(ean_p, []).append(p)
        for clave, etiqueta, data in cadenas:
            matches = 0
            for pc in data:
                ean = _ean_canonico(str(pc.get("ean", "")).strip())
                precio = pc.get("precio", 0)
                if not ean or precio <= 0:
                    continue
                for p in idx_ean.get(ean, []):
                    p["precios"][clave] = precio
                    p["fuentes"][clave] = {
                        "nombre":         pc.get("nombre", ""),
                        "imagen":         pc.get("imagen", ""),
                        "link":           pc.get("link", ""),
                        "fecha_scraping": pc.get("fecha_scraping", ""),
                        # gran parte de la gondola tiene oferta: precio =
                        # efectivo (lo que paga el publico hoy); regular +
                        # texto para mostrar ambos
                        "precio_regular": pc.get("precio_regular", 0),
                        "oferta":         pc.get("oferta", ""),
                    }
                    matches += 1
            if data:
                print(f"  {etiqueta}: {matches} productos del catalogo con precio gondola "
                      f"(de {len(data)} scrapeados; el resto es surtido exclusivo -> Fase B)")

    # Marcar precios stale. Umbral 14 dias: en Argentina, con inflacion, un precio
    # de hace 2 semanas ya no es confiable. El frontend usa este flag para avisar.
    STALE_DIAS = 14
    from datetime import date as _date
    hoy = _date.today()
    _mayoristas = ("yaguar", "maxicarrefour", "maxiconsumo", "nini", "coto", "carrefour", "dia", "masonline", "jumbo")
    precios_stale = 0
    for p in catalogo:
        fuentes = p.get("fuentes", {})
        for mayorista in _mayoristas:
            fuente = fuentes.get(mayorista, {})
            fecha_str = fuente.get("fecha_scraping", "")
            if fecha_str:
                try:
                    fecha = _date.fromisoformat(fecha_str)
                    dias = (hoy - fecha).days
                    fuente["dias_desde_scraping"] = dias
                    if dias > STALE_DIAS:
                        fuente["precio_stale"] = True
                        precios_stale += 1
                except ValueError:
                    pass
    if precios_stale:
        print(f"  Precios stale (>{STALE_DIAS} dias):     {precios_stale}")

    # Reporte de precios sospechosos para revisión manual (regla 08/09).
    # Lee las marcas que dejó la validación cruzada sobre el catálogo ya fusionado,
    # así el conteo refleja el estado final guardado.
    sospechosos = []
    for p in catalogo:
        for may, fuente in p.get("fuentes", {}).items():
            if fuente.get("precio_sospechoso"):
                otros = [p["precios"][m] for m in _mayoristas
                         if m != may and p["precios"].get(m, 0) > 0]
                ref = (sorted(otros)[len(otros) // 2]) if otros else 0
                sospechosos.append((p["nombre_display"], may, p["precios"].get(may, 0), ref))
    if sospechosos:
        sospechosos.sort(key=lambda x: (x[2] / x[3]) if x[3] else 0, reverse=True)
        print(f"\n  PRECIOS SOSPECHOSOS: {len(sospechosos)} productos (una fuente >2.5x la otra)")
        print(f"  Revisar contra la web del mayorista antes de confiar:")
        for nombre, may, precio, ref in sospechosos[:5]:
            print(f"    - {nombre[:45]}: {may} ${precio:.0f} vs otra ~${ref:.0f}")

    # Stats
    con_yag  = sum(1 for p in catalogo if p["precios"]["yaguar"] > 0)
    con_mc   = sum(1 for p in catalogo if p["precios"]["maxicarrefour"] > 0)
    con_mco  = sum(1 for p in catalogo if p["precios"]["maxiconsumo"] > 0)
    con_nini = sum(1 for p in catalogo if p["precios"].get("nini", 0) > 0)
    con_coto = sum(1 for p in catalogo if p["precios"].get("coto", 0) > 0)
    con_carr = sum(1 for p in catalogo if p["precios"].get("carrefour", 0) > 0)
    con_dia  = sum(1 for p in catalogo if p["precios"].get("dia", 0) > 0)
    con_mas  = sum(1 for p in catalogo if p["precios"].get("masonline", 0) > 0)
    con_jum  = sum(1 for p in catalogo if p["precios"].get("jumbo", 0) > 0)
    # Las metricas de comparativa miden SOLO mayoristas (coto es referencia gondola)
    _KEYS_MAY = ("yaguar", "maxicarrefour", "maxiconsumo", "nini")
    def _n_mayoristas(p):
        return sum(1 for k in _KEYS_MAY if p["precios"].get(k, 0) > 0)
    multi    = sum(1 for p in catalogo if _n_mayoristas(p) >= 2)
    todos_may = sum(1 for p in catalogo if _n_mayoristas(p) == len(_KEYS_MAY))
    abc_a_multi = sum(1 for p in catalogo if p.get("abc") == "A" and _n_mayoristas(p) >= 2)
    sin_img  = sum(1 for p in catalogo if not p.get("imagen"))

    print(f"\n{'='*60}")
    print(f"RESULTADO FINAL")
    print(f"{'='*60}")
    print(f"  Total productos con precio:   {len(catalogo)}")
    print(f"  Con precio Yaguar:            {con_yag}")
    print(f"  Con precio MaxiCarrefour:     {con_mc}")
    print(f"  Con precio Maxiconsumo:       {con_mco}")
    print(f"  Con precio Nini:              {con_nini}")
    print(f"  Con precio Coto (gondola):    {con_coto}")
    print(f"  Con precio Carrefour (gond.): {con_carr}")
    print(f"  Con precio Dia (gondola):     {con_dia}")
    print(f"  Con precio Masonline (gond.): {con_mas}")
    print(f"  Con precio Jumbo (gondola):   {con_jum}")
    print(f"  Con 2+ precios (comparativa): {multi}")
    print(f"  Con los {len(_KEYS_MAY)} precios mayoristas: {todos_may}")
    print(f"  ABC=A con 2+ precios:         {abc_a_multi}")
    print(f"  Sin imagen:                   {sin_img}")

    # Normalizar link de Maxiconsumo: ficha directa sin el prefijo de sucursal (ver
    # _maxiconsumo_product_link).
    for _p in catalogo:
        _mco = _p.get("fuentes", {}).get("maxiconsumo")
        if _mco and _mco.get("link"):
            _mco["link"] = _maxiconsumo_product_link(_mco["link"])

    # Links Carrefour: verificar cada EAN y caer a busqueda por nombre si el buscador da 0.
    _carrefour_links_hibrido(catalogo)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False, indent=2)

    print(f"\n  Guardado en: {OUTPUT_FILE}")

    # ------------------------------------------------------------------
    # Auto-aprendizaje: guardar matches fuzzy de alta confianza (>= 0.85)
    # en mapeo_brujula.json para que la próxima corrida los use como Capa 0.
    # Condición extra: el EAN debe estar confirmado en MaxiCarrefour (precio > 0).
    # Nunca sobreescribe entradas existentes.
    # ------------------------------------------------------------------
    mapeo_brujula_file = os.path.join(RAW_DIR, "mapeo_brujula.json")
    if (aprendizaje_yag or aprendizaje_mco or aprendizaje_nini) and os.path.isfile(mapeo_brujula_file):
        with open(mapeo_brujula_file, encoding="utf-8") as f:
            mb = json.load(f)

        mc_eans = {
            str(p.get("ean", "")).strip()
            for p in maxicarre
            if p.get("ean") and p.get("precio", 0) > 0
        }

        # Rechazos manuales de Facu (matches_pendientes.json): tampoco este
        # mecanismo puede re-aprenderlos — sin esto, un mapeo eliminado a mano
        # volvia a entrar por el fuzzy 1b en la corrida siguiente
        _rechazados_mb = set()
        _mp_file = os.path.join(BASE_DIR, "data", "quality", "matches_pendientes.json")
        if os.path.isfile(_mp_file):
            try:
                with open(_mp_file, encoding="utf-8") as f:
                    _rechazados_mb = set(json.load(f).get("rechazados", []))
            except (json.JSONDecodeError, OSError):
                pass

        nuevas_yag = nuevas_mco = nuevas_nini = 0

        for sku, ean in aprendizaje_yag.items():
            if ean not in mc_eans or f"yaguar:{sku}:{ean}" in _rechazados_mb:
                continue
            if sku not in mb["por_sku_yaguar"]:
                mb["por_sku_yaguar"][sku] = ean
                nuevas_yag += 1
                if ean not in mb["por_ean"]:
                    mb["por_ean"][ean] = {
                        "sector": "", "subcategoria": "", "abc": "", "nombre_verificacion": ""
                    }

        for sku, ean in aprendizaje_mco.items():
            if ean not in mc_eans or f"maxiconsumo:{sku}:{ean}" in _rechazados_mb:
                continue
            if sku not in mb["por_sku_maxiconsumo"]:
                mb["por_sku_maxiconsumo"][sku] = ean
                nuevas_mco += 1
                if ean not in mb["por_ean"]:
                    mb["por_ean"][ean] = {
                        "sector": "", "subcategoria": "", "abc": "", "nombre_verificacion": ""
                    }

        mb.setdefault("por_sku_nini", {})
        for sku, ean in aprendizaje_nini.items():
            if ean not in mc_eans or f"nini:{sku}:{ean}" in _rechazados_mb:
                continue
            if sku not in mb["por_sku_nini"]:
                mb["por_sku_nini"][sku] = ean
                nuevas_nini += 1
                if ean not in mb["por_ean"]:
                    mb["por_ean"][ean] = {
                        "sector": "", "subcategoria": "", "abc": "", "nombre_verificacion": ""
                    }

        if nuevas_yag + nuevas_mco + nuevas_nini > 0:
            with open(mapeo_brujula_file, "w", encoding="utf-8") as f:
                json.dump(mb, f, ensure_ascii=False, indent=2)
            print(f"\nAprendizaje: +{nuevas_yag} SKUs Yaguar, +{nuevas_mco} SKUs Maxiconsumo, "
                  f"+{nuevas_nini} SKUs Nini nuevos al mapeo")
        else:
            print("\nAprendizaje: sin asociaciones nuevas esta corrida (todo ya estaba en el mapeo)")
    elif not os.path.isfile(mapeo_brujula_file):
        print("\nAprendizaje: mapeo_brujula.json no encontrado, saltando")

    print("=" * 60)


if __name__ == "__main__":
    main()
