#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Construye un maestro dinamico de EANs a partir de MaxiCarrefour (100% EAN).
Los EANs nuevos (no presentes en el Listado Maestro estatico) se indexan por
nombre normalizado para enriquecer Yaguar y Maxiconsumo en las siguientes corridas.

Guardar en: data/raw/maestro_dinamico.json
Ejecutar: antes de enriquecer_eans.py y actualizar_catalogo.py
"""

import os
import re
import json
import glob
import unicodedata

import openpyxl

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
RAW_DIR       = os.path.join(BASE_DIR, "data", "raw")
MAESTRO_FILE  = os.path.join(RAW_DIR, "Listado Maestro 09-03.xlsx")
MC_DIR        = os.path.join(BASE_DIR, "targets", "maxicarrefour")
OUTPUT_FILE   = os.path.join(RAW_DIR, "maestro_dinamico.json")


def clave_nombre(nombre):
    n = (nombre or "").lower().strip()
    n = unicodedata.normalize("NFD", n)
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    n = re.sub(r"(\d),(\d)", r"\1.\2", n)
    n = re.sub(r"[^a-z0-9. ]", " ", n)
    n = re.sub(r"\bx\s*(\d)", r"\1", n)
    n = re.sub(r"(\d+\.?\d*)\s*lts?\b", lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+\.?\d*)\s*lt\b",   lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+\.?\d*)\s*l\b",    lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+)\s*grs?\b",       lambda m: m.group(1)+"gr", n)
    n = re.sub(r"(\d+\.?\d*)\s*kgs?\b", lambda m: str(int(float(m.group(1))*1000))+"gr", n)
    n = re.sub(r"(\d+)\s*(cc|ml|gr|kg|un)\b", r"\1\2", n)
    n = re.sub(r"\.", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def cargar_eans_maestro_estatico():
    eans = set()
    if not os.path.isfile(MAESTRO_FILE):
        print(f"  [WARN] Maestro estatico no encontrado: {MAESTRO_FILE}")
        return eans
    wb = openpyxl.load_workbook(MAESTRO_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_ean = next((i for i, h in enumerate(headers) if "EAN" in h.upper() or "BARRAS" in h.upper()), None)
    if col_ean is None:
        wb.close()
        return eans
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = row[col_ean]
        if not raw:
            continue
        try:
            ean = str(int(float(str(raw))))
            if len(ean) >= 7:
                eans.add(ean)
        except (ValueError, TypeError):
            pass
    wb.close()
    print(f"  Maestro estatico: {len(eans)} EANs cargados")
    return eans


def encontrar_mc_mejor():
    """Usa el archivo con más productos (no el más reciente — puede tener precio=0)."""
    archivos = glob.glob(os.path.join(MC_DIR, "output_maxicarrefour_*.json"))
    if not archivos:
        return None
    mejor = max(archivos, key=lambda f: len(json.load(open(f, encoding="utf-8"))))
    return mejor


def main():
    print("=" * 55)
    print("CONSTRUIR MAESTRO DINAMICO")
    print("=" * 55)

    eans_estaticos = cargar_eans_maestro_estatico()

    mc_file = encontrar_mc_mejor()
    if not mc_file:
        print("  ERROR: No se encontro output de MaxiCarrefour")
        return

    print(f"  MaxiCarrefour: {os.path.basename(mc_file)}")
    with open(mc_file, encoding="utf-8") as f:
        mc_data = json.load(f)

    por_nombre = {}
    nuevos = 0
    ya_en_maestro = 0

    for p in mc_data:
        ean = str(p.get("ean", "") or "").strip()
        nombre = (p.get("nombre", "") or "").strip()
        precio = p.get("precio", 0)

        if not ean or not nombre:
            continue
        try:
            ean = str(int(float(ean)))
        except (ValueError, TypeError):
            continue
        if len(ean) < 7:
            continue

        if ean in eans_estaticos:
            ya_en_maestro += 1
            continue

        clave = clave_nombre(nombre)
        if not clave or len(clave) < 5:
            continue

        if clave not in por_nombre:
            por_nombre[clave] = ean
            nuevos += 1

    resultado = {"por_nombre": por_nombre}
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False)

    print(f"  EANs ya en maestro estatico: {ya_en_maestro}")
    print(f"  EANs nuevos indexados:        {nuevos}")
    print(f"  Guardado en: {OUTPUT_FILE}")
    print("=" * 55)


if __name__ == "__main__":
    main()
